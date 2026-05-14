"""
Formulasyon-Optimizer v1
Farmasötik İnhaler Formülasyon Geliştirme & Optimizasyon Aracı
PyQt6 + Matplotlib + scipy + statsmodels + pyDOE3
ADIM 1: Ana iskelet, tema, veri modeli, Sekme 1 (Deney Tasarımı)
"""
import sys, os, json, math, datetime, itertools
import numpy as np
import pandas as pd
from scipy import stats

import pyDOE3

import matplotlib
matplotlib.use("QtAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QSplitter,
    QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QLineEdit, QPushButton, QCheckBox, QComboBox,
    QTabWidget, QScrollArea, QFrame, QFileDialog,
    QMessageBox, QDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QSizePolicy, QSpacerItem, QGroupBox,
    QDialogButtonBox, QAbstractItemView, QSpinBox,
    QDoubleSpinBox, QTextEdit, QRadioButton, QButtonGroup,
    QStatusBar
)
from PyQt6.QtCore import Qt, QSize, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont, QIcon

# ─── resource_path ─────────────────────────────────────────────────────────────
def resource_path(rel):
    base = getattr(sys, '_MEIPASS',
        os.path.dirname(os.path.abspath(
            sys.executable if getattr(sys, 'frozen', False) else __file__)))
    return os.path.join(base, rel)

# ═══════════════════════════════════════════════════════════════════════════════
# RENKLER & STİL
# ═══════════════════════════════════════════════════════════════════════════════
BG    = "#0e1219"
BG2   = "#141824"
BG3   = "#1c2336"
NAVY  = "#002D62"
GOLD  = "#FFC600"
TXT   = "#e0eaf8"
TXT2  = "#7090b0"
GREEN = "#00B050"
RED   = "#C00000"
TEAL  = "#0a8a6a"
CP    = ["#2E75B6","#ED7D31","#70AD47","#E84040","#7030A0",
         "#00B0F0","#D4A000","#C00000","#00B050","#FF69B4"]

RESPONSE_LABELS = {
    "mmad":      "MMAD (µm)",
    "gsd":       "GSD",
    "fpd_5um":   "FPD <5µm (mg)",
    "fpf_5um":   "FPF <5µm (%)",
    "fpd_3um":   "FPD <3µm (mg)",
    "fpf_3um":   "FPF <3µm (%)",
    "fpd_15um":  "FPD <1.5µm (mg)",
    "fpf_15um":  "FPF <1.5µm (%)",
    "metered":   "Metered Doz (mg)",
    "delivered": "Delivered Doz (mg)",
}

DESIGN_TYPES = [
    "Full Factorial (2k)",
    "Fractional Factorial (2k-p)",
    "Central Composite (CCD/RSM)",
    "Box-Behnken (BBD)",
    "Plackett-Burman",
    "One Factor at a Time (OFAT)",
]

DESIGN_INFO = {
    "Full Factorial (2k)": {
        "aciklama": "Tüm faktörlerin tüm kombinasyonlarını test eder. 2 seviyeli (düşük/yüksek).",
        "ne_zaman": "Faktör sayısı ≤4 ve tüm etkileşimler önemliyse.",
        "avantaj":  "Tüm ana etkiler + etkileşimler hesaplanır. Sonuçlar kesin.",
        "dezavantaj": "Faktör sayısı arttıkça run sayısı katlanır (2⁵=32, 2⁶=64...).",
        "oneri":    "2-3 faktör için ideal başlangıç tasarımı.",
        "run_formula": lambda k: 2**k,
        "seviye": 2,
    },
    "Fractional Factorial (2k-p)": {
        "aciklama": "Full Factorial'ın bir kesriyle ana etkileri belirler.",
        "ne_zaman": "5+ faktörü hızlıca taramak istediğinde, hangilerinin önemli olduğu bilinmiyorsa.",
        "avantaj":  "Full Factorial'a göre çok daha az run. Yüksek faktör sayısında kullanışlı.",
        "dezavantaj": "Bazı etkileşimler hesaplanamaz (aliasing). Tarama için uygundur.",
        "oneri":    "İlk tarama → anlamlı faktörleri bul → CCD/BBD ile optimize et.",
        "run_formula": lambda k: max(8, 2**(k-1)),
        "seviye": 2,
    },
    "Central Composite (CCD/RSM)": {
        "aciklama": "Faktör sınırlarının dışına çıkan eksenel noktalar içeren 5-seviyeli tasarım.",
        "ne_zaman": "2-4 faktörün optimumunu bulmak istediğinde, eğrisel ilişki bekliyorsan.",
        "avantaj":  "5 seviye ile quadratic model kurulur. Gerçek optimumu yakalayabilir.",
        "dezavantaj": "α noktaları faktör sınırlarının dışına çıkabilir.",
        "oneri":    "Tarama sonrası anlamlı 2-3 faktörle kullan. En güçlü optimizasyon tasarımı.",
        "run_formula": lambda k: 2**k + 2*k + 4,
        "seviye": 5,
    },
    "Box-Behnken (BBD)": {
        "aciklama": "Faktör sınırları içinde kalan, köşe noktası olmayan 3-seviyeli tasarım.",
        "ne_zaman": "3-4 faktörün optimumu, faktör sınırlarının dışına çıkmak istemiyorsan.",
        "avantaj":  "CCD'ye göre az run, köşe noktaları yok — aşırı kombinasyonları test etmez.",
        "dezavantaj": "En az 3 faktör gerektirir. 2 faktörle kullanılamaz.",
        "oneri":    "Formülasyon kısıtı olan çalışmalarda CCD'ye tercih edilir.",
        "run_formula": lambda k: 2*k*(k-1) + 3 if k >= 3 else 0,
        "seviye": 3,
    },
    "Plackett-Burman": {
        "aciklama": "Minimum run sayısıyla maksimum faktörü tarayan 2-seviyeli tarama tasarımı.",
        "ne_zaman": "5+ faktörü minimum run ile taramak istediğinde.",
        "avantaj":  "Çok az run (N = 4'ün katı). 11 faktörü 12 run'da tarayabilirsin.",
        "dezavantaj": "Sadece ana etkiler hesaplanır, etkileşimler hesaplanamaz.",
        "oneri":    "İlk tarama çalışması için en verimli tasarım.",
        "run_formula": lambda k: max(8, ((k + 1) // 4 + 1) * 4),
        "seviye": 2,
    },
    "One Factor at a Time (OFAT)": {
        "aciklama": "Her seferinde tek faktörü değiştirir, diğerleri sabit kalır.",
        "ne_zaman": "Referans veya tek faktörü izole etmek istediğinde.",
        "avantaj":  "Anlaşılması ve uygulanması kolay.",
        "dezavantaj": "Faktörler arası etkileşimleri göremez.",
        "oneri":    "Karşılaştırma amaçlı kullan. Optimizasyon için tercih etme.",
        "run_formula": lambda k: 2*k + 1,
        "seviye": 2,
    },
}

STYLE = f"""
QMainWindow, QDialog {{
    background: {BG};
}}
QWidget {{
    background: {BG};
    color: {TXT};
    font-family: 'Segoe UI', 'Arial';
    font-size: 13px;
}}
QLabel {{ color: {TXT}; background: transparent; }}
QLineEdit, QDoubleSpinBox, QSpinBox, QTextEdit {{
    background: {BG3};
    border: 1px solid #2a4060;
    border-radius: 4px;
    padding: 3px 6px;
    color: {TXT};
}}
QLineEdit:focus, QDoubleSpinBox:focus, QSpinBox:focus {{ border: 1px solid {GOLD}; }}
QPushButton {{
    background: {BG3};
    border: 1px solid #2a4060;
    border-radius: 5px;
    padding: 5px 12px;
    color: {TXT};
    font-weight: 500;
}}
QPushButton:hover {{ background: #253a5e; border-color: {GOLD}; }}
QPushButton:pressed {{ background: {BG2}; }}
QPushButton:disabled {{ color: #3a5070; border-color: #1a2a40; }}
QTabWidget::pane {{
    border: 1px solid #2a4060;
    background: {BG2};
    border-radius: 4px;
}}
QTabBar::tab {{
    background: {BG3};
    color: {TXT2};
    border: 1px solid #2a4060;
    padding: 7px 18px;
    border-bottom: none;
    margin-right: 2px;
    border-radius: 4px 4px 0 0;
    font-size: 12px;
}}
QTabBar::tab:selected {{
    background: {BG2};
    color: {GOLD};
    border-color: {GOLD};
    font-weight: bold;
}}
QTabBar::tab:hover:!selected {{ background: #253a5e; color: {TXT}; }}
QComboBox {{
    background: {BG3};
    border: 1px solid #2a4060;
    border-radius: 4px;
    padding: 4px 8px;
    color: {TXT};
}}
QComboBox::drop-down {{ border: none; width: 20px; }}
QComboBox QAbstractItemView {{
    background: {BG2};
    color: {TXT};
    selection-background-color: #1F4E79;
}}
QScrollArea {{ border: none; background: transparent; }}
QScrollBar:vertical {{
    background: {BG2}; width: 8px; border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: #2a4060; border-radius: 4px; min-height: 20px;
}}
QScrollBar:horizontal {{
    background: {BG2}; height: 8px; border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: #2a4060; border-radius: 4px; min-width: 20px;
}}
QTableWidget {{
    background: {BG2};
    gridline-color: #2a4060;
    color: {TXT};
    border: 1px solid #2a4060;
    border-radius: 4px;
}}
QTableWidget::item {{ padding: 4px; }}
QTableWidget::item:selected {{ background: #1F4E79; }}
QHeaderView::section {{
    background: #1F4E79;
    color: white;
    padding: 5px;
    border: 1px solid #2a4060;
    font-weight: bold;
    font-size: 12px;
}}
QGroupBox {{
    border: 1px solid #2a4060;
    border-radius: 6px;
    margin-top: 10px;
    padding-top: 6px;
    color: {TXT2};
    font-size: 11px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 4px;
    color: {GOLD};
}}
QCheckBox {{ color: {TXT}; background: transparent; spacing: 6px; }}
QCheckBox::indicator {{
    width: 14px; height: 14px;
    border: 1px solid #2a4060;
    border-radius: 3px;
    background: {BG3};
}}
QCheckBox::indicator:checked {{
    background: #2E75B6;
    border-color: #4a95d6;
}}
QRadioButton {{ color: {TXT}; background: transparent; spacing: 6px; }}
QRadioButton::indicator {{
    width: 14px; height: 14px;
    border: 1px solid #2a4060;
    border-radius: 7px;
    background: {BG3};
}}
QRadioButton::indicator:checked {{
    background: {GOLD};
    border-color: {GOLD};
}}
QStatusBar {{
    background: {BG2};
    color: {TXT2};
    border-top: 1px solid #2a4060;
    font-size: 11px;
}}
"""

# ═══════════════════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════════════════
def make_btn(text, color=None, height=32):
    b = QPushButton(text)
    b.setFixedHeight(height)
    c = color or "rgba(30,50,90,0.8)"
    b.setStyleSheet(f"""
        QPushButton {{
            background: {c};
            border: 1px solid rgba(255,255,255,0.15);
            border-radius: 5px;
            padding: 4px 14px;
            color: {TXT};
            font-weight: 500;
        }}
        QPushButton:hover {{ background: rgba(255,255,255,0.08); border-color: {GOLD}; }}
        QPushButton:pressed {{ background: rgba(0,0,0,0.2); }}
        QPushButton:disabled {{ color: #3a5070; border-color: #1a2a40; background: rgba(20,30,50,0.5); }}
    """)
    return b

def section_label(text):
    lbl = QLabel(text)
    lbl.setStyleSheet(f"color: {GOLD}; font-size: 11px; font-weight: bold; background: transparent;")
    return lbl

def card_frame():
    f = QFrame()
    f.setStyleSheet(f"""
        QFrame {{
            background: {BG2};
            border: 1px solid #2a4060;
            border-radius: 8px;
        }}
        QLabel {{ background: transparent; }}
    """)
    return f

def info_label(text):
    lbl = QLabel(text)
    lbl.setStyleSheet(f"color: {TXT2}; font-size: 11px; background: transparent;")
    lbl.setWordWrap(True)
    return lbl

def make_help_btn(tooltip_text, parent=None):
    btn = QPushButton("?")
    btn.setFixedSize(18, 18)
    btn.setStyleSheet(f"""
        QPushButton {{
            background: rgba(30,60,100,0.8);
            border: 1px solid #4a7ab0;
            border-radius: 9px;
            color: #90c0f0;
            font-size: 10px;
            font-weight: bold;
            padding: 0px;
        }}
        QPushButton:hover {{
            background: rgba(50,100,160,0.9);
            border-color: {GOLD};
            color: {GOLD};
        }}
    """)
    btn.clicked.connect(lambda: QMessageBox.information(parent, "Açıklama", tooltip_text))
    return btn

def tr2ascii(text):
    replacements = {
        'ç':'c','ğ':'g','ı':'i','ö':'o','ş':'s','ü':'u',
        'Ç':'C','Ğ':'G','İ':'I','Ö':'O','Ş':'S','Ü':'U',
    }
    result = str(text)
    for tr_char, ascii_char in replacements.items():
        result = result.replace(tr_char, ascii_char)
    return result

# ═══════════════════════════════════════════════════════════════════════════════
# VERİ MODELİ — Tüm proje verisi burada tutulur
# ═══════════════════════════════════════════════════════════════════════════════
class OptimizerProject:
    """Tüm sekmelerin paylaştığı merkezi veri modeli."""
    def __init__(self):
        self.reset()

    def reset(self):
        # Faktörler: [{"name": str, "unit": str, "low": float, "mid": float, "high": float, "type": "continuous"}]
        self.factors        = []
        # Seçili yanıtlar: ["mmad", "fpf_5um", ...]
        self.responses      = []
        # Tasarım tipi
        self.design_type    = "Central Composite (CCD/RSM)"
        # Tasarım matrisi (DataFrame) — faktör değerleri
        self.design_matrix  = None
        # Run sonuçları: {run_idx: {"mmad": float, "fpf_5um": float, ...}}
        self.run_results    = {}
        # Model sonuçları: {resp_key: statsmodels_model}
        self.model_results  = {}
        self.model_errors   = {}
        # Optimizasyon sonucu
        self.opt_solutions  = []   # Top-5 liste
        # Spesifikasyon sınırları: {"mmad": {"lsl": float, "usl": float, "goal": str, "weight": float}}
        self.spec_limits    = {}
        # Doğrulama sonuçları
        self.validation_results = {}

    def get_safe_names(self):
        """Faktör isimlerini istatistiksel model için güvenli hale getirir."""
        safe = []
        for i, f in enumerate(self.factors):
            s = f["name"].replace(" ", "_").replace("-", "_")
            s = "".join(c if c.isalnum() or c == "_" else "_" for c in s)
            if not s or s[0].isdigit():
                s = f"F{i+1}_{s}"
            safe.append(s)
        return safe

    def get_coded_df(self):
        """Tasarım matrisini kodlanmış değerlerle DataFrame olarak döner."""
        if self.design_matrix is None:
            return None, None
        safe_names = self.get_safe_names()
        df = pd.DataFrame()
        for i, f in enumerate(self.factors):
            col_name = f["name"]
            if col_name not in self.design_matrix.columns:
                continue
            vals = self.design_matrix[col_name].values
            lo, hi = f["low"], f["high"]
            mid = f.get("mid", (lo + hi) / 2)
            coded = []
            for v in vals:
                if hi == lo:
                    coded.append(0.0)
                elif v <= mid and (mid - lo) != 0:
                    coded.append((v - lo) / (mid - lo) - 1)
                elif (hi - mid) != 0:
                    coded.append((v - mid) / (hi - mid))
                else:
                    coded.append(0.0)
            df[safe_names[i]] = coded
        return df, safe_names

    def build_response_table(self):
        """Run sonuçlarını DataFrame olarak döner — eksik değerler NaN."""
        if self.design_matrix is None:
            return None
        rows = []
        for i in range(len(self.design_matrix)):
            row = {}
            for resp in self.responses:
                val = self.run_results.get(i, {}).get(resp, None)
                try:
                    row[resp] = float(val) if val is not None else np.nan
                except (TypeError, ValueError):
                    row[resp] = np.nan
            rows.append(row)
        return pd.DataFrame(rows, columns=self.responses)

    def generate_design_matrix(self):
        """Seçilen tasarım tipine göre matris oluşturur."""
        k = len(self.factors)
        if k == 0:
            return None, "En az 1 faktör ekleyin."

        design_type = self.design_type
        info = DESIGN_INFO.get(design_type, {})

        try:
            if design_type == "Full Factorial (2k)":
                coded = pyDOE3.ff2n(k)

            elif design_type == "Fractional Factorial (2k-p)":
                coded = pyDOE3.fracfact(" ".join([chr(97+i) for i in range(k)]))

            elif design_type == "Central Composite (CCD/RSM)":
                coded = pyDOE3.ccdesign(k, center=(4, 4), face="circumscribed")

            elif design_type == "Box-Behnken (BBD)":
                if k < 3:
                    return None, "Box-Behnken en az 3 faktör gerektirir."
                coded = pyDOE3.bbdesign(k, center=3)

            elif design_type == "Plackett-Burman":
                n_runs = max(8, ((k + 1) // 4 + 1) * 4)
                coded = pyDOE3.pbdesign(n_runs)
                coded = coded[:, :k]

            elif design_type == "One Factor at a Time (OFAT)":
                rows = []
                center = np.zeros(k)
                rows.append(center.copy())
                for i in range(k):
                    low_row = center.copy(); low_row[i] = -1; rows.append(low_row)
                    high_row = center.copy(); high_row[i] = 1; rows.append(high_row)
                coded = np.array(rows)

            else:
                return None, f"Bilinmeyen tasarım tipi: {design_type}"

        except Exception as e:
            return None, f"Matris oluşturma hatası: {e}"

        # Kodlanmış → gerçek değerlere çevir
        df = pd.DataFrame()
        for i, f in enumerate(self.factors):
            lo, hi = f["low"], f["high"]
            mid = f.get("mid", (lo + hi) / 2)
            col = coded[:, i]
            real = []
            for v in col:
                if v < 0:
                    real.append(lo + (v + 1) * (mid - lo))
                elif v > 0:
                    real.append(mid + v * (hi - mid))
                else:
                    real.append(mid)
            df[f["name"]] = [round(x, 6) for x in real]

        self.design_matrix = df
        self.run_results = {}
        self.model_results = {}
        return df, None

# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 1 — DENEY TASARIMI
# ═══════════════════════════════════════════════════════════════════════════════
class FactorRow(QWidget):
    """Tek bir faktör satırı — isim, birim, düşük, orta, yüksek."""
    deleted = pyqtSignal(object)

    def __init__(self, idx, parent=None):
        super().__init__(parent)
        self.idx = idx
        self.setStyleSheet("background: transparent;")
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 2, 0, 2)
        lay.setSpacing(6)

        # Sıra numarası
        num = QLabel(f"{idx+1}.")
        num.setFixedWidth(22)
        num.setStyleSheet(f"color: {TXT2}; font-size: 11px; background: transparent;")
        lay.addWidget(num)

        # İsim
        self.name_edit = QLineEdit(f"Faktör {idx+1}")
        self.name_edit.setFixedWidth(140)
        self.name_edit.setPlaceholderText("Faktör adı")
        lay.addWidget(self.name_edit)

        # Birim
        self.unit_edit = QLineEdit("")
        self.unit_edit.setFixedWidth(70)
        self.unit_edit.setPlaceholderText("Birim")
        lay.addWidget(self.unit_edit)

        # Düşük — etiket yok, başlıkta yazıyor
        self.low_sp = QDoubleSpinBox()
        self.low_sp.setRange(-1e9, 1e9); self.low_sp.setValue(0); self.low_sp.setDecimals(4)
        self.low_sp.setFixedWidth(100)
        lay.addWidget(self.low_sp)

        lay.addSpacing(6)

        # Orta — etiket yok, başlıkta yazıyor
        self.lbl_mid = QLabel("")   # gizli yer tutucu (genişlik için)
        self.lbl_mid.setFixedWidth(0)
        self.lbl_mid.setVisible(False)
        lay.addWidget(self.lbl_mid)
        self.mid_sp = QDoubleSpinBox()
        self.mid_sp.setRange(-1e9, 1e9); self.mid_sp.setValue(0.5); self.mid_sp.setDecimals(4)
        self.mid_sp.setFixedWidth(100)
        lay.addWidget(self.mid_sp)

        lay.addSpacing(6)

        # Üst — etiket yok, başlıkta yazıyor
        self.high_sp = QDoubleSpinBox()
        self.high_sp.setRange(-1e9, 1e9); self.high_sp.setValue(1); self.high_sp.setDecimals(4)
        self.high_sp.setFixedWidth(100)
        lay.addWidget(self.high_sp)

        lay.addStretch()

        # Sil butonu
        btn_del = QPushButton("✕")
        btn_del.setFixedSize(24, 24)
        btn_del.setStyleSheet(f"""
            QPushButton {{
                background: rgba(160,30,30,0.5);
                border: 1px solid #5a1010;
                border-radius: 12px;
                color: #ff8080;
                font-size: 11px;
            }}
            QPushButton:hover {{ background: rgba(200,40,40,0.8); color: white; }}
        """)
        btn_del.clicked.connect(lambda: self.deleted.emit(self))
        lay.addWidget(btn_del)

    def get_factor(self):
        return {
            "name": self.name_edit.text().strip() or f"Faktör {self.idx+1}",
            "unit": self.unit_edit.text().strip(),
            "low":  self.low_sp.value(),
            "mid":  self.mid_sp.value(),
            "high": self.high_sp.value(),
            "type": "continuous",
        }


class Tab1_Design(QWidget):
    """Sekme 1 — Deney Tasarımı: faktör girişi + tasarım seçimi + matris önizleme."""
    design_generated = pyqtSignal()   # Diğer sekmeleri uyarmak için

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project  = project
        self.app      = app_ref
        self._factor_rows = []
        self._build()

    def _build(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(12)

        # ── Sol panel: Faktörler + Tasarım tipi ──────────────────────────────
        left = QVBoxLayout()
        left.setSpacing(10)

        # — Faktörler kartı —
        fcard = card_frame()
        fl = QVBoxLayout(fcard)
        fl.setContentsMargins(14, 12, 14, 12)
        fl.setSpacing(8)

        fhdr = QHBoxLayout()
        fhdr.addWidget(section_label("📋  Faktörler"))
        fhdr.addWidget(make_help_btn(
            "Her faktör için:\n"
            "• İsim: Faktörün adı (örn. Oleik Asit)\n"
            "• Birim: Ölçü birimi (örn. %)\n"
            "• Alt: Minimum test seviyesi\n"
            "• Orta: Merkez nokta (CCD/BBD için önemli)\n"
            "• Üst: Maksimum test seviyesi\n\n"
            "En az 2 faktör girilmesi önerilir.", self))
        fhdr.addStretch()
        btn_add = make_btn("+ Faktör Ekle", "rgba(20,70,20,0.8)", 28)
        btn_add.clicked.connect(self._add_factor)
        fhdr.addWidget(btn_add)
        fl.addLayout(fhdr)

        # Sütun başlıkları
        hdr_w = QWidget(); hdr_w.setStyleSheet("background:transparent;")
        hdr_l = QHBoxLayout(hdr_w)
        hdr_l.setContentsMargins(28, 0, 32, 0); hdr_l.setSpacing(12)
        # Başlık sütunları — FactorRow genişlikleriyle eşleşir
        _hdr_items = [
            ("İsim",        140),
            ("Birim",        70),
            ("Alt Seviye",  100),
            ("Orta Seviye", 100),
            ("Üst Seviye",  100),
        ]
        for txt, w in _hdr_items:
            l = QLabel(txt)
            l.setFixedWidth(w)
            l.setStyleSheet(
                f"color:{GOLD}; font-size:10px; font-weight:bold; background:transparent;")
            hdr_l.addWidget(l)
            if txt == "Orta Seviye":
                self._mid_hdr_lbl = l
        hdr_l.addStretch()
        fl.addWidget(hdr_w)

        # Faktör satırları scroll alanı
        self.factor_scroll = QScrollArea()
        self.factor_scroll.setWidgetResizable(True)
        self.factor_scroll.setFixedHeight(220)
        self.factor_scroll.setStyleSheet("border: none; background: transparent;")
        self.factor_container = QWidget()
        self.factor_container.setStyleSheet("background: transparent;")
        self.factor_layout = QVBoxLayout(self.factor_container)
        self.factor_layout.setSpacing(2)
        self.factor_layout.setContentsMargins(0, 0, 0, 0)
        self.factor_layout.addStretch()
        self.factor_scroll.setWidget(self.factor_container)
        fl.addWidget(self.factor_scroll)

        # Başlangıç faktörleri (2 tane)
        self._add_factor(name="Oleik Asit", unit="%", low=0.5, mid=1.0, high=2.0)
        self._add_factor(name="Tween 80",   unit="%", low=0.05, mid=0.1, high=0.2)

        left.addWidget(fcard)

        # — Tasarım tipi kartı —
        dcard = card_frame()
        dl = QVBoxLayout(dcard)
        dl.setContentsMargins(14, 12, 14, 12)
        dl.setSpacing(10)
        dl.addWidget(section_label("🔬  Tasarım Tipi"))

        dcontent = QHBoxLayout()
        dcontent.setSpacing(16)

        # Sol: seçim + run bilgisi
        dsel = QVBoxLayout()
        dsel.setSpacing(8)
        self.design_combo = QComboBox()
        self.design_combo.addItems(DESIGN_TYPES)
        self.design_combo.setCurrentText("Central Composite (CCD/RSM)")
        self.design_combo.setFixedHeight(32)
        self.design_combo.currentTextChanged.connect(self._on_design_changed)
        dsel.addWidget(self.design_combo)

        self.run_info_lbl = QLabel("")
        self.run_info_lbl.setStyleSheet(f"color:{GOLD}; font-size:13px; font-weight:bold; background:transparent;")
        dsel.addWidget(self.run_info_lbl)

        self.seviye_lbl = QLabel("")
        self.seviye_lbl.setStyleSheet(f"color:{TXT2}; font-size:11px; background:transparent;")
        dsel.addWidget(self.seviye_lbl)

        dsel.addStretch()
        dcontent.addLayout(dsel)

        # Sağ: tasarım açıklaması
        self.design_info_box = QFrame()
        self.design_info_box.setStyleSheet(f"""
            QFrame {{
                background: {BG3};
                border: 1px solid #1a3050;
                border-radius: 6px;
            }}
            QLabel {{ background: transparent; }}
        """)
        dib_l = QVBoxLayout(self.design_info_box)
        dib_l.setContentsMargins(10, 8, 10, 8)
        dib_l.setSpacing(4)
        self.lbl_ne_zaman = info_label("")
        self.lbl_avantaj  = info_label("")
        self.lbl_dezavantaj = info_label("")
        self.lbl_oneri    = info_label("")
        for w in [self.lbl_ne_zaman, self.lbl_avantaj, self.lbl_dezavantaj, self.lbl_oneri]:
            dib_l.addWidget(w)
        dcontent.addWidget(self.design_info_box, 1)

        dl.addLayout(dcontent)
        left.addWidget(dcard)

        # — Yanıt seçimi kartı —
        rcard = card_frame()
        rl = QVBoxLayout(rcard)
        rl.setContentsMargins(14, 10, 14, 10)
        rl.setSpacing(8)
        rl.addWidget(section_label("🎯  Yanıt Değişkenleri (CQA)"))

        rgrid = QGridLayout()
        rgrid.setSpacing(6)
        self.resp_checks = {}
        items = list(RESPONSE_LABELS.items())
        for i, (key, lbl) in enumerate(items):
            cb = QCheckBox(lbl)
            cb.setChecked(key in ("mmad", "fpf_5um"))
            self.resp_checks[key] = cb
            rgrid.addWidget(cb, i // 3, i % 3)
        rl.addLayout(rgrid)
        left.addWidget(rcard)

        left.addStretch()

        # — Oluştur butonu —
        btn_gen = make_btn("▶  Deney Matrisi Oluştur", "rgba(20,70,120,0.9)", 42)
        btn_gen.setStyleSheet(btn_gen.styleSheet() + f"font-size: 14px; font-weight: bold;")
        btn_gen.clicked.connect(self._generate)
        left.addWidget(btn_gen)

        left_widget = QWidget()
        left_widget.setLayout(left)
        left_widget.setMinimumWidth(520)

        # ── Sağ panel: Matris önizleme + özet ───────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(10)

        # Özet bilgi çubuğu
        self.summary_frame = card_frame()
        sf_l = QHBoxLayout(self.summary_frame)
        sf_l.setContentsMargins(14, 8, 14, 8)
        sf_l.setSpacing(30)
        self.lbl_s_design  = QLabel("Tasarım: —")
        self.lbl_s_runs    = QLabel("Run: —")
        self.lbl_s_factors = QLabel("Faktör: —")
        self.lbl_s_resps   = QLabel("Yanıt: —")
        for l in [self.lbl_s_design, self.lbl_s_runs, self.lbl_s_factors, self.lbl_s_resps]:
            l.setStyleSheet(f"color:{TXT}; font-size:12px; font-weight:bold; background:transparent;")
            sf_l.addWidget(l)
        sf_l.addStretch()
        right.addWidget(self.summary_frame)

        # Tablo başlığı + export butonu
        tbl_hdr = QHBoxLayout()
        tbl_hdr.addWidget(section_label("📊  Deney Matrisi Önizleme"))
        tbl_hdr.addStretch()
        self.btn_export = make_btn("⬇  Excel'e Aktar", "rgba(20,80,20,0.8)", 28)
        self.btn_export.clicked.connect(self._export_excel)
        self.btn_export.setEnabled(False)
        tbl_hdr.addWidget(self.btn_export)
        right.addLayout(tbl_hdr)

        self.matrix_table = QTableWidget()
        self.matrix_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.matrix_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.matrix_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.matrix_table.setMinimumHeight(300)
        right.addWidget(self.matrix_table, 1)

        # Uyarı / bilgi alanı
        self.info_txt = QTextEdit()
        self.info_txt.setReadOnly(True)
        self.info_txt.setFixedHeight(80)
        self.info_txt.setStyleSheet(f"""
            QTextEdit {{
                background: {BG3};
                border: 1px solid #1a3050;
                border-radius: 4px;
                color: {TXT2};
                font-size: 11px;
            }}
        """)
        self.info_txt.setPlaceholderText("Deney matrisi oluşturulduğunda bilgiler burada görünür...")
        right.addWidget(self.info_txt)

        right_widget = QWidget()
        right_widget.setLayout(right)

        outer.addWidget(left_widget)
        outer.addWidget(right_widget, 1)

        # İlk güncelleme
        self._on_design_changed(self.design_combo.currentText())

    # ── İç metodlar ────────────────────────────────────────────────────────────
    def _add_factor(self, name=None, unit="", low=0.0, mid=0.5, high=1.0):
        idx = len(self._factor_rows)
        row = FactorRow(idx, self)
        if name:
            row.name_edit.setText(name)
            row.unit_edit.setText(unit)
            row.low_sp.setValue(low)
            row.mid_sp.setValue(mid)
            row.high_sp.setValue(high)
        row.deleted.connect(self._remove_factor)
        self._factor_rows.append(row)
        self.factor_layout.insertWidget(self.factor_layout.count() - 1, row)
        self._update_run_estimate()

    def _remove_factor(self, row_widget):
        if len(self._factor_rows) <= 1:
            QMessageBox.warning(self, "", "En az 1 faktör olmalıdır.")
            return
        self._factor_rows.remove(row_widget)
        row_widget.setParent(None)
        # İndeks numaralarını güncelle
        for i, r in enumerate(self._factor_rows):
            r.idx = i
            r.findChild(QLabel).setText(f"{i+1}.")
        self._update_run_estimate()

    def _on_design_changed(self, design_type):
        self.project.design_type = design_type
        info = DESIGN_INFO.get(design_type, {})
        k = len(self._factor_rows)

        # Run tahmini
        try:
            n_runs = info["run_formula"](k) if info else 0
        except Exception:
            n_runs = 0
        self.run_info_lbl.setText(f"Tahmini Run Sayısı: {n_runs}")
        self.seviye_lbl.setText(f"Faktör Seviyesi: {info.get('seviye','?')}")

        # Açıklamalar
        self.lbl_ne_zaman.setText(f"🕐 Ne zaman: {info.get('ne_zaman','')}")
        self.lbl_avantaj.setText(f"✅ Avantaj: {info.get('avantaj','')}")
        self.lbl_dezavantaj.setText(f"⚠️  Dezavantaj: {info.get('dezavantaj','')}")
        self.lbl_oneri.setText(f"💡 Öneri: {info.get('oneri','')}")

        # Orta seviye — sadece CCD ve BBD'de göster
        needs_mid = design_type in (
            "Central Composite (CCD/RSM)", "Box-Behnken (BBD)")
        for row in self._factor_rows:
            row.mid_sp.setVisible(needs_mid)
            row.lbl_mid.setVisible(needs_mid)
        # Başlık sütununu da güncelle
        if hasattr(self, "_mid_hdr_lbl"):
            self._mid_hdr_lbl.setVisible(needs_mid)

    def _update_run_estimate(self):
        # design_combo henüz oluşturulmamış olabilir (_add_factor __init__'te çağrılır)
        if not hasattr(self, "design_combo") or not hasattr(self, "run_info_lbl"):
            return
        design_type = self.design_combo.currentText()
        info = DESIGN_INFO.get(design_type, {})
        k = len(self._factor_rows)
        try:
            n_runs = info["run_formula"](k) if info else 0
        except Exception:
            n_runs = 0
        self.run_info_lbl.setText(f"Tahmini Run Sayısı: {n_runs}")

    def _collect_factors(self):
        factors = []
        for row in self._factor_rows:
            f = row.get_factor()
            if f["low"] >= f["high"]:
                QMessageBox.warning(self, "Hata",
                    f"'{f['name']}': Alt seviye üst seviyeden küçük olmalıdır.")
                return None
            if not (f["low"] <= f["mid"] <= f["high"]):
                f["mid"] = (f["low"] + f["high"]) / 2
            factors.append(f)
        return factors

    def _generate(self):
        # Faktörleri topla
        factors = self._collect_factors()
        if factors is None:
            return

        # Yanıtları topla
        responses = [k for k, cb in self.resp_checks.items() if cb.isChecked()]
        if not responses:
            QMessageBox.warning(self, "", "En az 1 yanıt değişkeni seçin.")
            return

        # Projeye kaydet
        self.project.factors   = factors
        self.project.responses = responses
        self.project.design_type = self.design_combo.currentText()

        # Matris oluştur
        df, err = self.project.generate_design_matrix()
        if err:
            QMessageBox.critical(self, "Hata", err)
            return

        # Tabloyu doldur
        self._fill_table(df)

        # Özet güncelle
        k = len(factors)
        n = len(df)
        self.lbl_s_design.setText(f"Tasarım: {self.project.design_type}")
        self.lbl_s_runs.setText(f"Run: {n}")
        self.lbl_s_factors.setText(f"Faktör: {k}")
        self.lbl_s_resps.setText(f"Yanıt: {len(responses)}")

        # Bilgi mesajı
        info = DESIGN_INFO.get(self.project.design_type, {})
        resp_names = [RESPONSE_LABELS.get(r, r) for r in responses]
        msg = (
            f"✅ Deney matrisi oluşturuldu.\n"
            f"   {n} run  |  {k} faktör  |  {len(responses)} yanıt\n"
            f"   Yanıtlar: {', '.join(resp_names)}\n"
            f"   Sonraki adım: Veri Girişi sekmesinden NGI sonuçlarını girin."
        )
        self.info_txt.setText(msg)
        self.btn_export.setEnabled(True)

        # Diğer sekmeleri bilgilendir
        self.design_generated.emit()
        self.app.status_bar.showMessage(
            f"Deney matrisi oluşturuldu — {n} run, {k} faktör", 5000)

    def _fill_table(self, df):
        self.matrix_table.clear()
        if df is None or df.empty:
            return
        cols = ["Run"] + list(df.columns)
        self.matrix_table.setColumnCount(len(cols))
        self.matrix_table.setRowCount(len(df))
        self.matrix_table.setHorizontalHeaderLabels(cols)

        # Birim satırını başlığa ekle
        for ci, col in enumerate(df.columns):
            f = next((x for x in self.project.factors if x["name"] == col), None)
            unit = f["unit"] if f and f.get("unit") else ""
            hdr_text = f"{col}\n({unit})" if unit else col
            self.matrix_table.horizontalHeaderItem(ci + 1).setText(hdr_text)

        for ri, row in df.iterrows():
            # Run no
            item = QTableWidgetItem(str(ri + 1))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QColor("#1F4E79"))
            self.matrix_table.setItem(ri, 0, item)
            # Faktör değerleri
            for ci, col in enumerate(df.columns):
                val = row[col]
                item = QTableWidgetItem(f"{val:.4f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # Merkez nokta farklı renk
                f = next((x for x in self.project.factors if x["name"] == col), None)
                if f:
                    mid = f.get("mid", (f["low"] + f["high"]) / 2)
                    if abs(val - mid) < 1e-6:
                        item.setBackground(QColor("#1a3a1a"))
                self.matrix_table.setItem(ri, ci + 1, item)

        self.matrix_table.resizeColumnsToContents()

    def _export_excel(self):
        if self.project.design_matrix is None:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Aktar",
            f"DoE_Matrisi_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return
        try:
            df = self.project.design_matrix.copy()
            df.insert(0, "Run No", range(1, len(df) + 1))
            # Yanıt sütunlarını boş olarak ekle
            for resp in self.project.responses:
                df[RESPONSE_LABELS.get(resp, resp)] = ""
            df.to_excel(path, index=False, engine="openpyxl")
            QMessageBox.information(self, "Başarılı",
                f"Excel dosyası kaydedildi:\n{path}\n\n"
                f"NGI ölçümlerinizi sarı sütunlara girin.")
            self.app.status_bar.showMessage(f"Excel kaydedildi: {path}", 4000)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel kaydedilemedi:\n{e}")

    def refresh(self):
        """Diğer sekmelerden geri gelindiğinde güncelle."""
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 2 — VERİ GİRİŞİ
# ═══════════════════════════════════════════════════════════════════════════════
class Tab2_DataEntry(QWidget):
    """Her run için NGI ölçüm sonuçlarını giriş tablosu."""

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project = project
        self.app     = app_ref
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)

        # ── Üst bilgi çubuğu ────────────────────────────────────────────────
        info_card = card_frame()
        info_l = QHBoxLayout(info_card)
        info_l.setContentsMargins(14, 8, 14, 8)
        info_l.setSpacing(30)
        self.lbl_runs    = QLabel("Run: —")
        self.lbl_factors = QLabel("Faktör: —")
        self.lbl_resps   = QLabel("Yanıt: —")
        self.lbl_filled  = QLabel("Dolu: —")
        for l in [self.lbl_runs, self.lbl_factors, self.lbl_resps, self.lbl_filled]:
            l.setStyleSheet(f"color:{TXT}; font-size:12px; font-weight:bold; background:transparent;")
            info_l.addWidget(l)
        info_l.addStretch()

        self.btn_clear = make_btn("🗑  Temizle", "rgba(120,20,20,0.6)", 28)
        self.btn_clear.clicked.connect(self._clear_responses)
        info_l.addWidget(self.btn_clear)

        self.btn_import = make_btn("📂  Excel'den İçe Aktar", "rgba(20,60,100,0.8)", 28)
        self.btn_import.clicked.connect(self._import_excel)
        info_l.addWidget(self.btn_import)

        lay.addWidget(info_card)

        # ── Açıklama ────────────────────────────────────────────────────────
        hint = QLabel(
            "Her run için NGI ölçüm sonuçlarını girin.  "
            "Mavi sütunlar = faktör değerleri (salt okunur).  "
            "Yeşil sütunlar = NGI yanıtları (düzenlenebilir).  "
            "Boş bırakılan run'lar model kurulumuna dahil edilmez.")
        hint.setStyleSheet(f"color:{TXT2}; font-size:11px; background:transparent;")
        hint.setWordWrap(True)
        lay.addWidget(hint)

        # ── Ana tablo ───────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet(self.table.styleSheet() + """
            QTableWidget { alternate-background-color: #111620; }
        """)
        self.table.cellChanged.connect(self._on_cell_changed)
        lay.addWidget(self.table, 1)

        # ── Alt butonlar ────────────────────────────────────────────────────
        bot = QHBoxLayout()
        bot.setSpacing(10)

        self.lbl_validation = QLabel("")
        self.lbl_validation.setStyleSheet(f"color:{TXT2}; font-size:11px; background:transparent;")
        bot.addWidget(self.lbl_validation)
        bot.addStretch()

        self.btn_next = make_btn("Model Kur  ▶", "rgba(20,80,20,0.8)", 36)
        self.btn_next.setStyleSheet(self.btn_next.styleSheet() +
                                    "font-size:13px; font-weight:bold;")
        self.btn_next.clicked.connect(self._go_to_model)
        bot.addWidget(self.btn_next)
        lay.addLayout(bot)

    # ── Yardımcı ──────────────────────────────────────────────────────────────
    def _n_factor_cols(self):
        return len(self.project.factors)

    def _resp_col_indices(self):
        """Yanıt sütunlarının tablo indekslerini döner."""
        nf = self._n_factor_cols()
        return list(range(1 + nf, 1 + nf + len(self.project.responses)))

    # ── Public ────────────────────────────────────────────────────────────────
    def refresh(self):
        """Sekme 1'den gelindiğinde tabloyu yeniden oluştur."""
        if self.project.design_matrix is None:
            self._show_empty()
            return
        self._build_table()
        self._update_info()

    def _show_empty(self):
        self.table.clear()
        self.table.setRowCount(1)
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels([""])
        item = QTableWidgetItem(
            "Önce 'Deney Tasarımı' sekmesinden matris oluşturun.")
        item.setForeground(QColor(TXT2))
        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        item.setFlags(Qt.ItemFlag.ItemIsEnabled)
        self.table.setItem(0, 0, item)
        self.lbl_runs.setText("Run: —")
        self.lbl_factors.setText("Faktör: —")
        self.lbl_resps.setText("Yanıt: —")
        self.lbl_filled.setText("Dolu: —")

    def _build_table(self):
        """Design matrix + boş yanıt sütunlarıyla tabloyu oluştur."""
        self.table.blockSignals(True)
        df      = self.project.design_matrix
        factors = self.project.factors
        resps   = self.project.responses
        n_runs  = len(df)

        # Sütunlar: Run No | faktörler... | yanıtlar...
        headers = ["Run"] + \
                  [f"{f['name']}" + (f"\n({f['unit']})" if f.get("unit") else "")
                   for f in factors] + \
                  [RESPONSE_LABELS.get(r, r) for r in resps]

        self.table.setColumnCount(len(headers))
        self.table.setRowCount(n_runs)
        self.table.setHorizontalHeaderLabels(headers)

        nf = len(factors)

        for ri in range(n_runs):
            # Run no
            item = QTableWidgetItem(str(ri + 1))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QColor("#1F4E79"))
            item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            self.table.setItem(ri, 0, item)

            # Faktör değerleri — salt okunur, mavi
            for ci, f in enumerate(factors):
                val = df.iloc[ri][f["name"]]
                item = QTableWidgetItem(f"{val:.4f}")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setBackground(QColor("#0d1e36"))
                item.setForeground(QColor("#90b8e0"))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                self.table.setItem(ri, 1 + ci, item)

            # Yanıt sütunları — düzenlenebilir, yeşil tonu
            for ci, resp in enumerate(resps):
                col_idx = 1 + nf + ci
                existing = self.project.run_results.get(ri, {}).get(resp, "")
                val_str = f"{existing:.6g}" if isinstance(existing, float) else ""
                item = QTableWidgetItem(val_str)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setBackground(QColor("#0d2010"))
                item.setForeground(QColor(TXT))
                self.table.setItem(ri, col_idx, item)

        self.table.blockSignals(False)
        self._update_info()

    def _on_cell_changed(self, row, col):
        """Yanıt hücresi değiştiğinde project.run_results güncelle."""
        nf = self._n_factor_cols()
        resp_start = 1 + nf
        if col < resp_start:
            return
        resp_idx = col - resp_start
        if resp_idx >= len(self.project.responses):
            return
        resp_key = self.project.responses[resp_idx]
        item = self.table.item(row, col)
        if item is None:
            return
        txt = item.text().strip().replace(",", ".")
        if txt == "":
            # Boş — sil
            if row in self.project.run_results:
                self.project.run_results[row].pop(resp_key, None)
        else:
            try:
                val = float(txt)
                if row not in self.project.run_results:
                    self.project.run_results[row] = {}
                self.project.run_results[row][resp_key] = val
                # Hücre rengini dolu olarak işaretle
                item.setBackground(QColor("#0a2e14"))
                item.setForeground(QColor("#80e890"))
            except ValueError:
                item.setBackground(QColor("#3a0a0a"))
                item.setForeground(QColor("#ff6060"))
        self._update_info()

    def _update_info(self):
        df    = self.project.design_matrix
        resps = self.project.responses
        if df is None:
            return
        n_runs  = len(df)
        n_resps = len(resps)
        # Kaç run'da en az 1 yanıt dolu?
        filled = sum(1 for i in range(n_runs)
                     if any(isinstance(self.project.run_results.get(i, {}).get(r), float)
                            for r in resps))
        self.lbl_runs.setText(f"Run: {n_runs}")
        self.lbl_factors.setText(f"Faktör: {len(self.project.factors)}")
        self.lbl_resps.setText(f"Yanıt: {n_resps}")
        color = GREEN if filled == n_runs else GOLD if filled > 0 else TXT2
        self.lbl_filled.setStyleSheet(
            f"color:{color}; font-size:12px; font-weight:bold; background:transparent;")
        self.lbl_filled.setText(f"Dolu: {filled}/{n_runs}")

        # Doğrulama mesajı
        if filled == 0:
            msg = "Henüz veri girilmedi."
            self.lbl_validation.setStyleSheet(f"color:{TXT2}; font-size:11px; background:transparent;")
        elif filled < n_runs:
            eksik = n_runs - filled
            msg = f"⚠  {eksik} run eksik — model kurulabilir ama dikkatli değerlendirin."
            self.lbl_validation.setStyleSheet(f"color:{GOLD}; font-size:11px; background:transparent;")
        else:
            msg = f"✅  Tüm {n_runs} run tamamlandı — Model Kur'a geçebilirsiniz."
            self.lbl_validation.setStyleSheet(f"color:{GREEN}; font-size:11px; background:transparent;")
        self.lbl_validation.setText(msg)

    def _clear_responses(self):
        reply = QMessageBox.question(
            self, "Temizle",
            "Tüm yanıt değerleri silinecek. Devam edilsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.project.run_results = {}
            self.project.model_results = {}
            self._build_table()

    def _import_excel(self):
        """Excel dosyasından yanıt değerlerini içe aktar."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel Dosyası Seç", "", "Excel (*.xlsx *.xls)")
        if not path:
            return
        try:
            df_imp = pd.read_excel(path, engine="openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya okunamadı:\n{e}")
            return

        # Yanıt sütunlarını eşleştir
        imported = 0
        resps = self.project.responses
        resp_labels = {RESPONSE_LABELS.get(r, r): r for r in resps}

        for ri in range(min(len(df_imp), len(self.project.design_matrix))):
            for col_name, resp_key in resp_labels.items():
                # Hem tam isim hem kısmi eşleşme dene
                matched_col = None
                for c in df_imp.columns:
                    if col_name.lower() in str(c).lower() or str(c).lower() in col_name.lower():
                        matched_col = c
                        break
                if matched_col is None:
                    continue
                val = df_imp.iloc[ri].get(matched_col, None)
                if pd.notna(val):
                    try:
                        fval = float(val)
                        if ri not in self.project.run_results:
                            self.project.run_results[ri] = {}
                        self.project.run_results[ri][resp_key] = fval
                        imported += 1
                    except (ValueError, TypeError):
                        pass

        self._build_table()
        QMessageBox.information(self, "İçe Aktarma",
            f"{imported} değer başarıyla içe aktarıldı.")
        self.app.status_bar.showMessage(f"{imported} değer içe aktarıldı.", 4000)

    def _go_to_model(self):
        """Sekme 3'e geç."""
        dm = self.project.design_matrix
        if dm is None:
            QMessageBox.warning(self, "", "Önce Deney Tasarımı sekmesinden matris oluşturun.")
            return
        filled = sum(1 for i in range(len(dm))
                     if self.project.run_results.get(i))
        if filled == 0:
            QMessageBox.warning(self, "",
                "Model kurmak için en az birkaç run verisi gereklidir.")
            return
        self.app.tabs.setCurrentIndex(2)


# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 3 — MODEL
# ═══════════════════════════════════════════════════════════════════════════════
class ModelWorker(QThread):
    """OLS model fit işlemini arka planda çalıştırır."""
    finished = pyqtSignal(dict, dict)   # results, errors
    progress = pyqtSignal(str)

    def __init__(self, project):
        super().__init__()
        self.project = project

    def run(self):
        try:
            self._run_safe()
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            write_log(f"ModelWorker crash:\n{tb}")
            self.finished.emit({}, {"_global": f"Model crash: {e}\n\n{tb}"})

    def _run_safe(self):
        import statsmodels.api as sm
        from statsmodels.stats.anova import anova_lm
        import statsmodels.formula.api as smf

        results = {}
        errors  = {}
        p       = self.project
        safe    = p.get_safe_names()

        # Yanıt verisini topla
        resp_df = p.build_response_table()
        if resp_df is None:
            errors["_global"] = "Yanıt verisi bulunamadı."
            self.finished.emit(results, errors)
            return
        if len(resp_df) == 0:
            errors["_global"] = "Yanıt tablosu boş."
            self.finished.emit(results, errors)
            return

        # Kodlanmış faktör matrisi
        coded_df, _ = p.get_coded_df()
        if coded_df is None:
            errors["_global"] = "Tasarım matrisi bulunamadı."
            self.finished.emit(results, errors)
            return

        # coded_df index'ini sıfırla
        coded_df = coded_df.reset_index(drop=True)

        # Her yanıt için model kur
        for resp_key in p.responses:
            self.progress.emit(f"Model kuruluyor: {RESPONSE_LABELS.get(resp_key, resp_key)}")
            try:
                # Sütun yoksa atla
                if resp_key not in resp_df.columns:
                    errors[resp_key] = "Yanıt sütunu bulunamadı."
                    continue

                y = np.array(resp_df[resp_key].tolist(), dtype=float)

                # Eksik satırları çıkar (NaN olanlar)
                mask = ~np.isnan(y)
                mask_idx = np.where(mask)[0]
                min_obs = len(safe) + 2
                if int(mask.sum()) < min_obs:
                    errors[resp_key] = (
                        f"Yeterli veri yok: {int(mask.sum())} run dolu, "
                        f"en az {min_obs} gerekli.")
                    continue

                X_df    = coded_df.iloc[mask_idx].copy().reset_index(drop=True)
                y_clean = y[mask_idx]
                y_s     = pd.Series(y_clean, name=resp_key)
                n_factors = len(safe)

                # Karesel terimler
                for name in safe:
                    X_df[f"{name}_sq"] = X_df[name] ** 2

                # İkili etkileşimler
                for i in range(n_factors):
                    for j in range(i+1, n_factors):
                        col = f"{safe[i]}_x_{safe[j]}"
                        X_df[col] = X_df[safe[i]] * X_df[safe[j]]

                X_mat = sm.add_constant(X_df, has_constant="add")
                model = sm.OLS(y_s, X_mat).fit()

                # ANOVA — önce typ=1 dene, başarısız olursa manuel hesapla
                anova = None
                try:
                    from statsmodels.stats.anova import anova_lm as _anova_lm
                    anova = _anova_lm(model, typ=1)
                except Exception as _e1:
                    try:
                        ss_res = float(np.sum(model.resid**2))
                        ss_tot = float(np.sum((y_s - float(y_s.mean()))**2))
                        ss_reg = ss_tot - ss_res
                        df_reg = int(model.df_model)
                        df_res = int(model.df_resid)
                        ms_reg = ss_reg / max(df_reg, 1)
                        ms_res = ss_res / max(df_res, 1)
                        f_val  = ms_reg / max(ms_res, 1e-10)
                        from scipy.stats import f as _f_dist
                        p_f    = float(1 - _f_dist.cdf(f_val, df_reg, df_res))
                        anova  = pd.DataFrame({
                            "sum_sq": [ss_reg, ss_res],
                            "df":     [float(df_reg), float(df_res)],
                            "PR(>F)": [p_f, float("nan")],
                        }, index=["Model", "Artik"])
                    except Exception as _e2:
                        write_log(f"ANOVA hesaplanamadi ({resp_key}): {_e1} | {_e2}")

                # Artık normallik testi (Shapiro-Wilk 3-5000 arası çalışır)
                sw_p = None
                try:
                    n_res = len(model.resid)
                    if 3 <= n_res <= 5000:
                        _, sw_p = stats.shapiro(model.resid)
                except Exception:
                    pass

                results[resp_key] = {
                    "model":      model,
                    "anova":      anova,
                    "X_df":       X_df,
                    "y":          y_clean,
                    "sw_p":       sw_p,
                    "r2":         model.rsquared,
                    "r2_adj":     model.rsquared_adj,
                    "rmse":       float(np.sqrt(model.mse_resid)),
                    "f_pvalue":   float(model.f_pvalue),
                    "safe_names": safe,
                    "n_obs":      int(mask.sum()),
                }

            except Exception as e:
                import traceback
                errors[resp_key] = f"{e}\n{traceback.format_exc()}"

        self.finished.emit(results, errors)


class Tab3_Model(QWidget):
    """Sekme 3 — OLS model fit, R², ANOVA, katsayı tablosu, artık grafikleri."""

    model_ready = pyqtSignal()

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project = project
        self.app     = app_ref
        self._worker = None
        self._build()

    def _build(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(12)

        # ── Sol: kontroller + özet tablo ──────────────────────────────────────
        left = QVBoxLayout()
        left.setSpacing(10)

        # Başlık + buton
        top = QHBoxLayout()
        top.addWidget(section_label("📈  Regresyon Modeli"))
        top.addStretch()
        self.btn_fit = make_btn("▶  Model Kur", "rgba(20,80,20,0.9)", 34)
        self.btn_fit.setStyleSheet(self.btn_fit.styleSheet() +
                                   "font-size:13px; font-weight:bold;")
        self.btn_fit.clicked.connect(self._fit_models)
        top.addWidget(self.btn_fit)
        left.addLayout(top)

        # Model tipi açıklaması
        info = info_label(
            "Quadratic (ikinci dereceden) model: ana etkiler + ikili etkileşimler + "
            "karesel terimler. Seçilen yanıt değişkenlerine ayrı ayrı OLS modeli kurulur.")
        left.addWidget(info)

        # Yanıt seçim combo
        resp_row = QHBoxLayout()
        resp_row.addWidget(QLabel("Yanıt:"))
        self.resp_combo = QComboBox()
        self.resp_combo.setFixedHeight(28)
        self.resp_combo.currentTextChanged.connect(self._on_resp_changed)
        resp_row.addWidget(self.resp_combo)
        resp_row.addStretch()
        left.addLayout(resp_row)

        # ── Özet kart ─────────────────────────────────────────────────────────
        sum_card = card_frame()
        sl = QGridLayout(sum_card)
        sl.setContentsMargins(14, 10, 14, 10)
        sl.setSpacing(8)

        def stat_pair(label, attr):
            lbl = QLabel(label)
            lbl.setStyleSheet(f"color:{TXT2}; font-size:11px; background:transparent;")
            val = QLabel("—")
            val.setStyleSheet(f"color:{GOLD}; font-size:14px; font-weight:bold; background:transparent;")
            setattr(self, attr, val)
            return lbl, val

        for row, (lbl_txt, attr) in enumerate([
            ("R²",           "lbl_r2"),
            ("Adj R²",       "lbl_r2adj"),
            ("RMSE",         "lbl_rmse"),
            ("Model p",      "lbl_fp"),
            ("Shapiro-Wilk p", "lbl_swp"),
            ("Gözlem (n)",   "lbl_nobs"),
        ]):
            lbl, val = stat_pair(lbl_txt, attr)
            sl.addWidget(lbl, row, 0)
            sl.addWidget(val, row, 1)

        left.addWidget(sum_card)

        # ── Katsayı tablosu ───────────────────────────────────────────────────
        left.addWidget(section_label("🔢  Katsayılar & Anlamlılık"))
        self.coef_table = QTableWidget()
        self.coef_table.setColumnCount(5)
        self.coef_table.setHorizontalHeaderLabels(
            ["Terim", "Katsayı", "Std Hata", "t", "p"])
        self.coef_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents)
        self.coef_table.horizontalHeader().setStretchLastSection(True)
        self.coef_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.coef_table.setMinimumHeight(200)
        left.addWidget(self.coef_table, 1)

        # ── ANOVA tablosu ─────────────────────────────────────────────────────
        left.addWidget(section_label("📊  ANOVA Tablosu"))
        self.anova_table = QTableWidget()
        self.anova_table.setColumnCount(5)
        self.anova_table.setHorizontalHeaderLabels(
            ["Kaynak", "SS", "df", "MS", "p"])
        self.anova_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents)
        self.anova_table.horizontalHeader().setStretchLastSection(True)
        self.anova_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.anova_table.setFixedHeight(180)
        left.addWidget(self.anova_table)

        left_w = QWidget()
        left_w.setLayout(left)
        left_w.setMinimumWidth(480)
        left_w.setMaximumWidth(560)

        # ── Sağ: Artık grafikleri ─────────────────────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(8)
        right.addWidget(section_label("🔬  Artık (Residual) Analizi"))

        # 4 grafik: Predicted vs Actual, Residuals vs Fitted,
        #           Normal Q-Q, Residuals vs Order
        self.fig = Figure(figsize=(9, 7), facecolor=BG2)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setStyleSheet(f"background: {BG2};")
        right.addWidget(self.canvas, 1)

        # Durum mesajı
        self.status_lbl = QLabel("Model henüz kurulmadı.")
        self.status_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        right.addWidget(self.status_lbl)

        # İleri butonu
        bot = QHBoxLayout()
        bot.addStretch()
        self.btn_next = make_btn("Response Surface  ▶", "rgba(20,80,20,0.8)", 36)
        self.btn_next.setStyleSheet(self.btn_next.styleSheet() +
                                    "font-size:13px; font-weight:bold;")
        self.btn_next.setEnabled(False)
        self.btn_next.clicked.connect(lambda: self.app.tabs.setCurrentIndex(3))
        bot.addWidget(self.btn_next)
        right.addLayout(bot)

        right_w = QWidget()
        right_w.setLayout(right)

        outer.addWidget(left_w)
        outer.addWidget(right_w, 1)

    # ── Public ────────────────────────────────────────────────────────────────
    def refresh(self):
        if self.project.design_matrix is None:
            return
        # Yanıt combo'yu güncelle
        self.resp_combo.blockSignals(True)
        self.resp_combo.clear()
        for r in self.project.responses:
            self.resp_combo.addItem(RESPONSE_LABELS.get(r, r), r)
        self.resp_combo.blockSignals(False)
        # Daha önce model kurulmuşsa sonuçları göster
        if self.project.model_results:
            self._show_results_for_current()

    # ── Model fit ─────────────────────────────────────────────────────────────
    def _fit_models(self):
        if self.project.design_matrix is None:
            QMessageBox.warning(self, "", "Önce Deney Tasarımı sekmesinden matris oluşturun.")
            return
        filled = sum(1 for i in range(len(self.project.design_matrix))
                     if self.project.run_results.get(i))
        if filled == 0:
            QMessageBox.warning(self, "", "Önce Veri Girişi sekmesinden NGI sonuçlarını girin.")
            return

        self.btn_fit.setEnabled(False)
        self.btn_fit.setText("⏳ Hesaplanıyor...")
        self.status_lbl.setText("Model kuruluyor...")

        self._worker = ModelWorker(self.project)
        self._worker.finished.connect(self._on_model_done)
        self._worker.progress.connect(
            lambda msg: self.app.status_bar.showMessage(msg))
        self._worker.start()

    def _on_model_done(self, results, errors):
        self.btn_fit.setEnabled(True)
        self.btn_fit.setText("▶  Model Kur")

        if "_global" in errors:
            QMessageBox.critical(self, "Hata", errors["_global"])
            return

        self.project.model_results = results
        self.project.model_errors  = errors

        # Combo güncelle
        self.resp_combo.blockSignals(True)
        self.resp_combo.clear()
        for r in self.project.responses:
            label = RESPONSE_LABELS.get(r, r)
            if r in errors:
                label += "  ⚠"
            self.resp_combo.addItem(label, r)
        self.resp_combo.blockSignals(False)

        self._show_results_for_current()
        self.btn_next.setEnabled(bool(results))
        self.model_ready.emit()
        self.app.status_bar.showMessage(
            f"Model hazır — {len(results)} yanıt başarıyla modellendi.", 6000)

        if errors:
            err_list = "\n".join(f"  • {RESPONSE_LABELS.get(k,k)}: {v}"
                                 for k, v in errors.items() if k != "_global")
            QMessageBox.warning(self, "Uyarı",
                f"Bazı yanıtlar modellenemedi:\n{err_list}")

    def _on_resp_changed(self, _):
        self._show_results_for_current()

    def _show_results_for_current(self):
        resp_key = self.resp_combo.currentData()
        if resp_key is None:
            return
        res = self.project.model_results.get(resp_key)
        if res is None:
            err = self.project.model_errors.get(resp_key, "Model kurulmadı.")
            self.status_lbl.setText(f"⚠  {err}")
            self._clear_display()
            return

        model = res["model"]

        # ── Özet istatistikler ────────────────────────────────────────────────
        def fmt_p(p):
            if p is None: return "—"
            if p < 0.001: return "<0.001 ✅"
            if p < 0.05:  return f"{p:.4f} ✅"
            return f"{p:.4f} ⚠"

        def color_r2(v):
            c = GREEN if v >= 0.9 else GOLD if v >= 0.7 else RED
            return f'<span style="color:{c}">{v:.4f}</span>'

        self.lbl_r2.setText(f"{res['r2']:.4f}")
        self.lbl_r2.setStyleSheet(
            f"color:{'#00B050' if res['r2']>=0.9 else '#FFC600' if res['r2']>=0.7 else '#C00000'};"
            f"font-size:14px;font-weight:bold;background:transparent;")
        self.lbl_r2adj.setText(f"{res['r2_adj']:.4f}")
        self.lbl_rmse.setText(f"{res['rmse']:.4f}")
        self.lbl_fp.setText(fmt_p(res['f_pvalue']))
        sw_p = res.get("sw_p")
        self.lbl_swp.setText(
            f"{sw_p:.4f} {'✅ Normal' if sw_p and sw_p>0.05 else '⚠ Normal değil'}"
            if sw_p else "—")
        self.lbl_nobs.setText(str(res["n_obs"]))

        # ── Katsayı tablosu ───────────────────────────────────────────────────
        params  = model.params
        bse     = model.bse
        tvalues = model.tvalues
        pvalues = model.pvalues

        self.coef_table.setRowCount(len(params))
        for i, (name, coef) in enumerate(params.items()):
            p_val = pvalues[name]
            significant = p_val < 0.05

            name_item = QTableWidgetItem(name)
            if significant:
                name_item.setForeground(QColor(GREEN))
                name_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            else:
                name_item.setForeground(QColor(TXT2))

            items = [
                name_item,
                QTableWidgetItem(f"{coef:.4f}"),
                QTableWidgetItem(f"{bse[name]:.4f}"),
                QTableWidgetItem(f"{tvalues[name]:.3f}"),
                QTableWidgetItem(f"{p_val:.4f}"),
            ]
            for j, item in enumerate(items):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if j > 0:
                    if significant:
                        item.setBackground(QColor("#0a2a0a"))
                    else:
                        item.setBackground(QColor(BG3))
                self.coef_table.setItem(i, j, item)

            # p-değeri renklendirme
            p_item = self.coef_table.item(i, 4)
            if p_val < 0.001:
                p_item.setForeground(QColor(GREEN))
            elif p_val < 0.05:
                p_item.setForeground(QColor("#90e890"))
            elif p_val < 0.1:
                p_item.setForeground(QColor(GOLD))
            else:
                p_item.setForeground(QColor(TXT2))

        # ── ANOVA tablosu ─────────────────────────────────────────────────────
        anova = res.get("anova")
        if anova is not None:
            self.anova_table.setRowCount(len(anova))
            for i, (src, row) in enumerate(anova.iterrows()):
                p_v = row.get("PR(>F)", np.nan)
                items = [
                    QTableWidgetItem(str(src)),
                    QTableWidgetItem(f"{row.get('sum_sq', 0):.4f}"),
                    QTableWidgetItem(f"{int(row.get('df', 0))}"),
                    QTableWidgetItem(f"{row.get('sum_sq', 0)/max(row.get('df',1),1):.4f}"),
                    QTableWidgetItem(f"{p_v:.4f}" if not np.isnan(p_v) else "—"),
                ]
                for j, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(QColor(BG3 if i % 2 == 0 else BG2))
                    if j == 4 and not np.isnan(p_v):
                        item.setForeground(
                            QColor(GREEN if p_v < 0.05 else TXT2))
                    self.anova_table.setItem(i, j, item)

        # ── Artık grafikleri ──────────────────────────────────────────────────
        self._plot_residuals(res)

        self.status_lbl.setText(
            f"✅  Model hazır  |  R²={res['r2']:.4f}  |  Adj R²={res['r2_adj']:.4f}  "
            f"|  n={res['n_obs']}")

    def _clear_display(self):
        self.coef_table.setRowCount(0)
        self.anova_table.setRowCount(0)
        self.fig.clear()
        self.canvas.draw()
        for attr in ["lbl_r2","lbl_r2adj","lbl_rmse","lbl_fp","lbl_swp","lbl_nobs"]:
            getattr(self, attr).setText("—")

    def _plot_residuals(self, res):
        model = res["model"]
        fitted   = model.fittedvalues.values
        residuals = model.resid.values
        y_actual  = res["y"]

        self.fig.clear()
        self.fig.patch.set_facecolor(BG2)

        axes = self.fig.subplots(2, 2)
        plot_cfg = dict(facecolor=BG3, labelcolor=TXT, titlecolor=GOLD)

        for ax in axes.flat:
            ax.set_facecolor(BG3)
            ax.tick_params(colors=TXT2, labelsize=8)
            ax.xaxis.label.set_color(TXT2)
            ax.yaxis.label.set_color(TXT2)
            for spine in ax.spines.values():
                spine.set_edgecolor("#2a4060")

        # 1. Predicted vs Actual
        ax1 = axes[0, 0]
        mn = min(min(y_actual), min(fitted))
        mx = max(max(y_actual), max(fitted))
        ax1.scatter(y_actual, fitted, color=CP[0], s=40, zorder=3, alpha=0.85)
        ax1.plot([mn, mx], [mn, mx], '--', color=GOLD, lw=1, label="İdeal")
        ax1.set_xlabel("Gerçek"); ax1.set_ylabel("Tahmin")
        ax1.set_title("Tahmin vs Gerçek", color=GOLD, fontsize=9)
        ax1.legend(fontsize=7, labelcolor=TXT2,
                   facecolor=BG2, edgecolor="#2a4060")

        # 2. Residuals vs Fitted
        ax2 = axes[0, 1]
        ax2.scatter(fitted, residuals, color=CP[1], s=40, zorder=3, alpha=0.85)
        ax2.axhline(0, color=GOLD, lw=1, linestyle="--")
        ax2.set_xlabel("Tahmin"); ax2.set_ylabel("Artık")
        ax2.set_title("Artık vs Tahmin", color=GOLD, fontsize=9)

        # 3. Normal Q-Q
        ax3 = axes[1, 0]
        (osm, osr), (slope, intercept, _) = stats.probplot(residuals)
        ax3.scatter(osm, osr, color=CP[2], s=30, zorder=3, alpha=0.85)
        line_x = np.array([osm[0], osm[-1]])
        ax3.plot(line_x, slope * line_x + intercept,
                 '--', color=GOLD, lw=1)
        ax3.set_xlabel("Teorik Kantil"); ax3.set_ylabel("Örnek Kantil")
        ax3.set_title("Normal Q-Q", color=GOLD, fontsize=9)

        # 4. Residuals vs Order
        ax4 = axes[1, 1]
        ax4.plot(range(1, len(residuals)+1), residuals,
                 'o-', color=CP[3], markersize=4, lw=1, alpha=0.85)
        ax4.axhline(0, color=GOLD, lw=1, linestyle="--")
        ax4.set_xlabel("Gözlem Sırası"); ax4.set_ylabel("Artık")
        ax4.set_title("Artık vs Sıra", color=GOLD, fontsize=9)

        self.fig.tight_layout(pad=1.5)
        self.canvas.draw()


# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 4 — RESPONSE SURFACE
# ═══════════════════════════════════════════════════════════════════════════════
class Tab4_ResponseSurface(QWidget):
    """3D yüzey + kontur haritası + sabit faktör slider'ları."""

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project    = project
        self.app        = app_ref
        self._ds_worker = None
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(10)

        # ── Kontrol çubuğu ────────────────────────────────────────────────────
        ctrl = card_frame()
        cl = QHBoxLayout(ctrl)
        cl.setContentsMargins(14, 8, 14, 8)
        cl.setSpacing(16)

        cl.addWidget(QLabel("Yanıt:"))
        self.resp_combo = QComboBox(); self.resp_combo.setFixedHeight(28)
        cl.addWidget(self.resp_combo)

        cl.addWidget(QLabel("X Ekseni:"))
        self.x_combo = QComboBox(); self.x_combo.setFixedHeight(28)
        cl.addWidget(self.x_combo)

        cl.addWidget(QLabel("Y Ekseni:"))
        self.y_combo = QComboBox(); self.y_combo.setFixedHeight(28)
        cl.addWidget(self.y_combo)

        cl.addWidget(QLabel("Çözünürlük:"))
        self.res_spin = QSpinBox()
        self.res_spin.setRange(20, 100); self.res_spin.setValue(40)
        self.res_spin.setFixedWidth(60); self.res_spin.setFixedHeight(28)
        cl.addWidget(self.res_spin)

        cl.addStretch()

        self.btn_plot = make_btn("▶  Grafik Çiz", "rgba(20,80,20,0.9)", 32)
        self.btn_plot.setStyleSheet(self.btn_plot.styleSheet() +
                                    "font-size:12px; font-weight:bold;")
        self.btn_plot.clicked.connect(self._plot)
        cl.addWidget(self.btn_plot)

        outer.addWidget(ctrl)

        # ── Sabit faktör slider'ları ───────────────────────────────────────────
        self.slider_card = card_frame()
        self.slider_layout = QHBoxLayout(self.slider_card)
        self.slider_layout.setContentsMargins(14, 8, 14, 8)
        self.slider_layout.setSpacing(20)
        self.slider_card.setFixedHeight(70)
        self._slider_widgets = {}   # {factor_name: (label, spinbox)}
        outer.addWidget(self.slider_card)

        # ── Ana grafik alanı ─────────────────────────────────────────────────
        self.fig = Figure(figsize=(13, 5), facecolor=BG2)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setStyleSheet(f"background: {BG2};")
        outer.addWidget(self.canvas, 1)

        # ── Alt çubuk ────────────────────────────────────────────────────────
        bot = QHBoxLayout()
        self.status_lbl = QLabel("Model kurulduktan sonra grafik çizebilirsiniz.")
        self.status_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        bot.addWidget(self.status_lbl)
        bot.addStretch()
        self.btn_reset_view = make_btn("↺  Görünümü Sıfırla", "rgba(30,50,90,0.8)", 32)
        self.btn_reset_view.setToolTip("3D grafiği başlangıç açısına döndür")
        self.btn_reset_view.clicked.connect(self._reset_3d_view)
        bot.addWidget(self.btn_reset_view)
        self.btn_next = make_btn("Optimizasyon  ▶", "rgba(20,80,20,0.8)", 34)
        self.btn_next.setStyleSheet(self.btn_next.styleSheet() +
                                    "font-size:12px; font-weight:bold;")
        self.btn_next.clicked.connect(lambda: self.app.tabs.setCurrentIndex(4))
        bot.addWidget(self.btn_next)
        outer.addLayout(bot)

        # Sinyal bağlantıları
        self.resp_combo.currentTextChanged.connect(self._on_selection_changed)
        self.x_combo.currentTextChanged.connect(self._on_selection_changed)
        self.y_combo.currentTextChanged.connect(self._on_selection_changed)

    # ── Public ───────────────────────────────────────────────────────────────
    def refresh(self):
        if not self.project.model_results:
            return
        self._populate_combos()

    # ── İç metodlar ──────────────────────────────────────────────────────────
    def _populate_combos(self):
        """Combo kutularını model sonuçlarına göre doldur."""
        self.resp_combo.blockSignals(True)
        self.x_combo.blockSignals(True)
        self.y_combo.blockSignals(True)

        self.resp_combo.clear()
        for r in self.project.responses:
            if r in self.project.model_results:
                self.resp_combo.addItem(RESPONSE_LABELS.get(r, r), r)

        self.x_combo.clear()
        self.y_combo.clear()
        safe = self.project.get_safe_names()
        factors = self.project.factors
        for i, f in enumerate(factors):
            label = f"{f['name']}" + (f" ({f['unit']})" if f.get("unit") else "")
            self.x_combo.addItem(label, i)
            self.y_combo.addItem(label, i)

        # Varsayılan: X=0, Y=1
        if self.x_combo.count() >= 2:
            self.y_combo.setCurrentIndex(1)

        self.resp_combo.blockSignals(False)
        self.x_combo.blockSignals(False)
        self.y_combo.blockSignals(False)

        self._build_sliders()

    def _build_sliders(self):
        """Sabit faktörler için slider (SpinBox) oluştur."""
        # Mevcut widget'ları temizle
        while self.slider_layout.count():
            item = self.slider_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._slider_widgets = {}

        xi = self.x_combo.currentData() or 0
        yi = self.y_combo.currentData() or 1

        for i, f in enumerate(self.project.factors):
            if i in (xi, yi):
                continue
            mid = f.get("mid", (f["low"] + f["high"]) / 2)
            grp = QGroupBox(f"{f['name']}")
            grp.setFixedWidth(150)
            gl = QVBoxLayout(grp)
            gl.setContentsMargins(6, 4, 6, 4)
            sp = QDoubleSpinBox()
            sp.setRange(f["low"], f["high"])
            sp.setValue(mid)
            sp.setDecimals(3)
            sp.setSingleStep((f["high"] - f["low"]) / 20)
            sp.setFixedHeight(26)
            if f.get("unit"):
                sp.setSuffix(f"  {f['unit']}")
            gl.addWidget(sp)
            self.slider_layout.addWidget(grp)
            self._slider_widgets[i] = sp

        self.slider_layout.addStretch()

    def _on_selection_changed(self, _=None):
        self._build_sliders()

    def _get_fixed_values(self):
        """Sabit faktörlerin şu anki değerlerini döner: {factor_idx: value}"""
        return {i: sp.value() for i, sp in self._slider_widgets.items()}

    def _predict_grid(self, resp_key, xi, yi, n):
        """X-Y ızgara üzerinde tahmin yüzeyi hesapla."""
        res   = self.project.model_results[resp_key]
        model = res["model"]
        safe  = self.project.get_safe_names()
        factors = self.project.factors

        fx = factors[xi]; fy = factors[yi]
        x_real = np.linspace(fx["low"], fx["high"], n)
        y_real = np.linspace(fy["low"], fy["high"], n)
        XX, YY = np.meshgrid(x_real, y_real)

        fixed = self._get_fixed_values()

        # Kodlanmış değerlere çevir
        def encode(val, f):
            lo, hi = f["low"], f["high"]
            mid = f.get("mid", (lo + hi) / 2)
            if hi == lo: return 0.0
            if val <= mid and (mid - lo) != 0:
                return (val - lo) / (mid - lo) - 1
            elif (hi - mid) != 0:
                return (val - mid) / (hi - mid)
            return 0.0

        ZZ = np.zeros_like(XX)
        n_factors = len(factors)

        import statsmodels.api as _sm
        param_idx = model.params.index

        for r in range(n):
            for c in range(n):
                row_vals = {}
                for k, f in enumerate(factors):
                    if k == xi:
                        row_vals[safe[k]] = encode(XX[r, c], f)
                    elif k == yi:
                        row_vals[safe[k]] = encode(YY[r, c], f)
                    else:
                        row_vals[safe[k]] = encode(
                            fixed.get(k, f.get("mid", (f["low"]+f["high"])/2)), f)

                # Karesel ve etkileşim terimleri
                for name in safe:
                    row_vals[f"{name}_sq"] = row_vals[name] ** 2
                for a in range(n_factors):
                    for b in range(a+1, n_factors):
                        col = f"{safe[a]}_x_{safe[b]}"
                        row_vals[col] = row_vals[safe[a]] * row_vals[safe[b]]

                row_df  = pd.DataFrame([row_vals])
                row_mat = _sm.add_constant(row_df, has_constant="add")
                # Modelin beklediği sütunlarla hizala
                for col in param_idx:
                    if col not in row_mat.columns:
                        row_mat[col] = 0.0
                row_mat = row_mat.reindex(columns=param_idx, fill_value=0.0)
                try:
                    ZZ[r, c] = float(model.predict(row_mat)[0])
                except Exception:
                    ZZ[r, c] = np.nan

        return XX, YY, ZZ, x_real, y_real

    def _plot(self):
        if not self.project.model_results:
            QMessageBox.warning(self, "", "Önce Model sekmesinden model kurun.")
            return

        resp_key = self.resp_combo.currentData()
        xi = self.x_combo.currentData()
        yi = self.y_combo.currentData()

        if resp_key is None or xi is None or yi is None:
            return
        if xi == yi:
            QMessageBox.warning(self, "", "X ve Y ekseni farklı faktörler olmalıdır.")
            return
        if resp_key not in self.project.model_results:
            QMessageBox.warning(self, "", "Bu yanıt için model bulunamadı.")
            return

        self.btn_plot.setEnabled(False)
        self.btn_plot.setText("⏳ Hesaplanıyor...")
        self.app.status_bar.showMessage("Yüzey hesaplanıyor...")
        QApplication.processEvents()

        try:
            n   = self.res_spin.value()
            XX, YY, ZZ, x_real, y_real = self._predict_grid(resp_key, xi, yi, n)
            self._draw(resp_key, xi, yi, XX, YY, ZZ)
            self.status_lbl.setText(
                f"✅  {RESPONSE_LABELS.get(resp_key,resp_key)}  |  "
                f"X: {self.project.factors[xi]['name']}  |  "
                f"Y: {self.project.factors[yi]['name']}")
            self.app.status_bar.showMessage("Grafik hazır.", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Grafik çizilemedi:\n{e}")
        finally:
            self.btn_plot.setEnabled(True)
            self.btn_plot.setText("▶  Grafik Çiz")

    def _reset_3d_view(self):
        """3D yüzey grafiğini varsayılan açıya döndür."""
        try:
            for ax in self.fig.get_axes():
                if hasattr(ax, 'elev'):
                    ax.view_init(elev=30, azim=-60)
            self.canvas.draw()
        except Exception:
            pass

    def _draw(self, resp_key, xi, yi, XX, YY, ZZ):
        from mpl_toolkits.mplot3d import Axes3D
        import matplotlib.cm as cm

        factors = self.project.factors
        fx = factors[xi]; fy = factors[yi]
        resp_label = RESPONSE_LABELS.get(resp_key, resp_key)

        self.fig.clear()
        self.fig.patch.set_facecolor(BG2)

        # Sol: 3D yüzey
        ax3d = self.fig.add_subplot(1, 2, 1, projection='3d')
        ax3d.set_facecolor(BG3)
        ax3d.tick_params(colors=TXT2, labelsize=7)
        ax3d.xaxis.pane.fill = False; ax3d.yaxis.pane.fill = False
        ax3d.zaxis.pane.fill = False
        ax3d.xaxis.pane.set_edgecolor("#2a4060")
        ax3d.yaxis.pane.set_edgecolor("#2a4060")
        ax3d.zaxis.pane.set_edgecolor("#2a4060")

        surf = ax3d.plot_surface(XX, YY, ZZ,
                                  cmap="viridis", alpha=0.85,
                                  linewidth=0, antialiased=True)
        ax3d.set_xlabel(f"{fx['name']}" + (f"\n({fx['unit']})" if fx.get("unit") else ""),
                        color=TXT2, fontsize=8, labelpad=8)
        ax3d.set_ylabel(f"{fy['name']}" + (f"\n({fy['unit']})" if fy.get("unit") else ""),
                        color=TXT2, fontsize=8, labelpad=8)
        ax3d.set_zlabel(resp_label, color=TXT2, fontsize=8, labelpad=8)
        ax3d.set_title("3D Yüzey", color=GOLD, fontsize=10, pad=10)

        cb3d = self.fig.colorbar(surf, ax=ax3d, shrink=0.5, pad=0.1)
        cb3d.ax.tick_params(colors=TXT2, labelsize=7)

        # Sağ: kontur haritası
        ax2d = self.fig.add_subplot(1, 2, 2)
        ax2d.set_facecolor(BG3)
        ax2d.tick_params(colors=TXT2, labelsize=8)
        for spine in ax2d.spines.values():
            spine.set_edgecolor("#2a4060")

        levels = 20
        cf = ax2d.contourf(XX, YY, ZZ, levels=levels, cmap="viridis", alpha=0.9)
        cs = ax2d.contour(XX, YY, ZZ, levels=levels, colors="white",
                           alpha=0.25, linewidths=0.5)
        ax2d.clabel(cs, inline=True, fontsize=6, colors="white", fmt="%.2f")

        cb2d = self.fig.colorbar(cf, ax=ax2d)
        cb2d.ax.tick_params(colors=TXT2, labelsize=7)
        cb2d.set_label(resp_label, color=TXT2, fontsize=8)

        # Optimum noktayı işaretle
        if self.project.opt_solutions:
            best = self.project.opt_solutions[0]
            ox = best.get("factors", {}).get(self.project.factors[xi]["name"])
            oy = best.get("factors", {}).get(self.project.factors[yi]["name"])
            if ox is not None and oy is not None:
                ax2d.plot(ox, oy, "*", color=GOLD, markersize=14,
                          zorder=5, label="Optimum")
                ax2d.legend(fontsize=8, labelcolor=TXT2,
                             facecolor=BG2, edgecolor="#2a4060")

        # Gerçek ölçüm noktaları
        res = self.project.model_results.get(resp_key, {})
        y_data = res.get("y")
        if y_data is not None and self.project.design_matrix is not None:
            df = self.project.design_matrix
            xs_real, ys_real = [], []
            for ri in range(len(df)):
                if self.project.run_results.get(ri, {}).get(resp_key) is not None:
                    xs_real.append(df.iloc[ri][fx["name"]])
                    ys_real.append(df.iloc[ri][fy["name"]])
            if xs_real:
                ax2d.scatter(xs_real, ys_real, c="white", s=30,
                             zorder=4, edgecolors="#2a4060", linewidths=0.5,
                             alpha=0.8, label="Ölçüm")

        ax2d.set_xlabel(f"{fx['name']}" + (f" ({fx['unit']})" if fx.get("unit") else ""),
                        color=TXT2, fontsize=9)
        ax2d.set_ylabel(f"{fy['name']}" + (f" ({fy['unit']})" if fy.get("unit") else ""),
                        color=TXT2, fontsize=9)
        ax2d.set_title("Kontur Haritası", color=GOLD, fontsize=10)

        self.fig.tight_layout(pad=2.0)
        self.canvas.draw()


# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 5 — OPTİMİZASYON
# ═══════════════════════════════════════════════════════════════════════════════
class OptWorker(QThread):
    """Derringer-Suich optimizasyonunu arka planda çalıştırır."""
    finished = pyqtSignal(list, str)   # solutions, error_msg

    def __init__(self, project, spec_limits):
        super().__init__()
        self.project     = project
        self.spec_limits = spec_limits

    def run(self):
        try:
            self._run_safe()
        except Exception as e:
            import traceback
            write_log(f"OptWorker crash:\n{traceback.format_exc()}")
            self.finished.emit([], str(e))

    def _run_safe(self):
        from scipy.optimize import differential_evolution
        import statsmodels.api as sm

        p      = self.project
        safe   = p.get_safe_names()
        factors = p.factors
        n_factors = len(factors)
        specs  = self.spec_limits

        def encode(val, f):
            lo, hi = f["low"], f["high"]
            mid = f.get("mid", (lo+hi)/2)
            if hi == lo: return 0.0
            if val <= mid and (mid-lo) != 0:
                return (val-lo)/(mid-lo) - 1
            elif (hi-mid) != 0:
                return (val-mid)/(hi-mid)
            return 0.0

        def predict_resp(resp_key, factor_vals):
            res = p.model_results.get(resp_key)
            if res is None:
                return None
            model = res["model"]
            row = {}
            for i, f in enumerate(factors):
                row[safe[i]] = encode(factor_vals[i], f)
            for name in safe:
                row[f"{name}_sq"] = row[name]**2
            for a in range(n_factors):
                for b in range(a+1, n_factors):
                    row[f"{safe[a]}_x_{safe[b]}"] = row[safe[a]]*row[safe[b]]
            df_row = pd.DataFrame([row])
            row_mat = sm.add_constant(df_row, has_constant="add")
            for col in model.params.index:
                if col not in row_mat.columns:
                    row_mat[col] = 0.0
            row_mat = row_mat.reindex(columns=model.params.index, fill_value=0.0)
            try:
                return float(model.predict(row_mat)[0])
            except Exception:
                return None

        def derringer_suich(pred, spec):
            """Gerçek Derringer-Suich desirability fonksiyonu."""
            goal = spec.get("goal", "target")
            lsl  = spec.get("lsl")
            usl  = spec.get("usl")
            tgt  = spec.get("target")
            s    = max(0.01, float(spec.get("weight", 1.0)))

            if pred is None or np.isnan(pred):
                return 0.0

            if goal == "maximize":
                if lsl is None: return 1.0
                if pred <= lsl: return 0.0
                if usl is not None and pred >= usl: return 1.0
                hi = usl if usl is not None else lsl + abs(lsl)*2 + 1
                return float(((pred - lsl)/(hi - lsl))**s)

            elif goal == "minimize":
                if usl is None: return 1.0
                if pred >= usl: return 0.0
                if lsl is not None and pred <= lsl: return 1.0
                lo = lsl if lsl is not None else usl - abs(usl)*2 - 1
                return float(((usl - pred)/(usl - lo))**s)

            else:  # target
                if tgt is None:
                    tgt = ((lsl or 0) + (usl or 0)) / 2
                if lsl is not None and pred < lsl: return 0.0
                if usl is not None and pred > usl: return 0.0
                if pred == tgt: return 1.0
                if pred < tgt and lsl is not None and tgt > lsl:
                    return float(((pred - lsl)/(tgt - lsl))**s)
                if pred > tgt and usl is not None and usl > tgt:
                    return float(((usl - pred)/(usl - tgt))**s)
                return 1.0

        active_resps = [r for r in p.responses
                        if r in p.model_results and r in specs]

        def neg_desirability(x):
            if not active_resps:
                return 0.0
            ds = []
            for resp_key in active_resps:
                pred = predict_resp(resp_key, x)
                d    = derringer_suich(pred, specs[resp_key])
                ds.append(d)
            # Geometrik ortalama
            prod = 1.0
            for d in ds:
                prod *= max(d, 1e-10)
            return -float(prod ** (1.0/len(ds)))

        bounds = [(f["low"], f["high"]) for f in factors]

        # differential_evolution — çoklu başlangıç noktası
        solutions = []
        seeds = [42, 7, 123, 999, 31]
        for seed in seeds:
            try:
                res = differential_evolution(
                    neg_desirability, bounds,
                    seed=seed, maxiter=200, tol=1e-6,
                    popsize=12, mutation=(0.5, 1.5), recombination=0.9,
                    workers=1, polish=True)
                if res.success or res.fun < 0:
                    d_score = -res.fun
                    preds = {}
                    for resp_key in p.responses:
                        if resp_key in p.model_results:
                            pv = predict_resp(resp_key, res.x)
                            preds[resp_key] = pv
                    factor_dict = {f["name"]: float(res.x[i])
                                   for i, f in enumerate(factors)}
                    solutions.append({
                        "desirability": round(d_score, 6),
                        "factors":      factor_dict,
                        "predictions":  preds,
                        "x":            list(res.x),
                    })
            except Exception:
                pass

        # Tekrarları ve benzer noktaları filtrele
        unique = []
        for s in sorted(solutions, key=lambda x: -x["desirability"]):
            is_dup = False
            for u in unique:
                dist = np.sqrt(sum((s["x"][i]-u["x"][i])**2
                                   for i in range(len(s["x"]))))
                span = np.sqrt(sum((f["high"]-f["low"])**2
                                   for f in factors)) or 1
                if dist/span < 0.05:
                    is_dup = True
                    break
            if not is_dup:
                unique.append(s)
            if len(unique) >= 5:
                break

        self.finished.emit(unique, "")


class Tab5_Optimization(QWidget):
    """Sekme 5 — Derringer-Suich optimizasyon, USL/LSL, Top-5 çözüm."""

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project    = project
        self.app        = app_ref
        self._worker    = None
        self._spec_rows = {}   # {resp_key: widget_dict}
        self._build()

    def _build(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(12)

        # ── Sol: Spesifikasyon tanımı ─────────────────────────────────────────
        left = QVBoxLayout()
        left.setSpacing(10)
        left.addWidget(section_label("🎯  CQA Hedefleri & Spesifikasyon Sınırları"))

        hint = info_label(
            "Her yanıt için hedef tipi, LSL/USL sınırları ve ağırlık (s) tanımlayın. "
            "s=1 lineer, s<1 daha toleranslı, s>1 daha katı. "
            "Boş bırakılan sınırlar göz ardı edilir.")
        left.addWidget(hint)

        # Spec tablosu başlıkları
        hdr = QWidget(); hdr.setStyleSheet("background:transparent;")
        hl  = QHBoxLayout(hdr); hl.setContentsMargins(0,0,0,0); hl.setSpacing(4)
        for txt, w in [("Yanıt", 130), ("Hedef", 90), ("LSL", 72),
                       ("USL", 72), ("Hedef Değer", 80), ("Ağırlık s", 72)]:
            l = QLabel(txt)
            l.setFixedWidth(w)
            l.setStyleSheet(f"color:{GOLD}; font-size:10px; font-weight:bold; background:transparent;")
            hl.addWidget(l)
        left.addWidget(hdr)

        # Spec satırları scroll alanı
        self.spec_scroll = QScrollArea()
        self.spec_scroll.setWidgetResizable(True)
        self.spec_scroll.setFixedHeight(280)
        self.spec_scroll.setStyleSheet("border:none; background:transparent;")
        self.spec_container = QWidget()
        self.spec_container.setStyleSheet("background:transparent;")
        self.spec_vlay = QVBoxLayout(self.spec_container)
        self.spec_vlay.setSpacing(4)
        self.spec_vlay.setContentsMargins(0,0,0,0)
        self.spec_vlay.addStretch()
        self.spec_scroll.setWidget(self.spec_container)
        left.addWidget(self.spec_scroll)

        # Optimize butonu
        self.btn_opt = make_btn("▶  Optimize Et", "rgba(20,80,20,0.9)", 42)
        self.btn_opt.setStyleSheet(self.btn_opt.styleSheet() +
                                   "font-size:14px; font-weight:bold;")
        self.btn_opt.clicked.connect(self._run_optimization)
        left.addWidget(self.btn_opt)

        # Robustluk özeti
        left.addWidget(section_label("🔩  Robustluk Özeti (Optimum ±10%)"))
        self.robust_table = QTableWidget()
        self.robust_table.setColumnCount(4)
        self.robust_table.setHorizontalHeaderLabels(
            ["Yanıt", "Optimum", "-10%", "+10%"])
        self.robust_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.robust_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.robust_table.setFixedHeight(180)
        left.addWidget(self.robust_table)

        left.addStretch()

        left_w = QWidget(); left_w.setLayout(left)
        left_w.setMinimumWidth(500); left_w.setMaximumWidth(560)

        # ── Sağ: Top-5 sonuçlar ───────────────────────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(8)
        right.addWidget(section_label("🏆  Top-5 Optimum Formülasyon"))

        self.result_tabs = QTabWidget()
        self.result_tabs.setStyleSheet(f"""
            QTabBar::tab {{ padding: 5px 12px; font-size: 11px; }}
        """)
        right.addWidget(self.result_tabs, 1)

        # Desirability grafiği
        right.addWidget(section_label("📊  Desirability Skoru"))
        self.fig_d = Figure(figsize=(8, 2.2), facecolor=BG2)
        self.canvas_d = FigureCanvas(self.fig_d)
        self.canvas_d.setFixedHeight(160)
        self.canvas_d.setStyleSheet(f"background:{BG2};")
        right.addWidget(self.canvas_d)

        # Alt buton
        bot = QHBoxLayout()
        self.status_lbl = QLabel("Model kurulduktan sonra optimize edebilirsiniz.")
        self.status_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        bot.addWidget(self.status_lbl)
        bot.addStretch()
        self.btn_next = make_btn("Design Space  ▶", "rgba(20,80,20,0.8)", 34)
        self.btn_next.setStyleSheet(self.btn_next.styleSheet() +
                                    "font-size:12px; font-weight:bold;")
        self.btn_next.clicked.connect(lambda: self.app.tabs.setCurrentIndex(5))
        bot.addWidget(self.btn_next)
        right.addLayout(bot)

        right_w = QWidget(); right_w.setLayout(right)

        outer.addWidget(left_w)
        outer.addWidget(right_w, 1)

    # ── Public ───────────────────────────────────────────────────────────────
    def refresh(self):
        if not self.project.model_results:
            return
        self._build_spec_rows()
        if self.project.opt_solutions:
            self._show_solutions(self.project.opt_solutions)

    # ── Spec satırları ────────────────────────────────────────────────────────
    def _build_spec_rows(self):
        while self.spec_vlay.count() > 1:
            item = self.spec_vlay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._spec_rows = {}

        GOAL_OPTS = ["maximize", "minimize", "target"]
        GOAL_DEFAULTS = {
            "mmad":     ("target",   "2.5", "3.5", "3.0"),
            "gsd":      ("minimize", "1.0", "2.5", ""),
            "fpd_5um":  ("maximize", "0.5", "",    ""),
            "fpf_5um":  ("maximize", "35",  "",    ""),
            "fpd_3um":  ("maximize", "0.3", "",    ""),
            "fpf_3um":  ("maximize", "25",  "",    ""),
            "fpd_15um": ("maximize", "0.1", "",    ""),
            "fpf_15um": ("maximize", "15",  "",    ""),
            "metered":  ("target",   "0.9", "1.1", "1.0"),
            "delivered":("maximize", "0.8", "",    ""),
        }

        for resp_key in self.project.responses:
            if resp_key not in self.project.model_results:
                continue
            label = RESPONSE_LABELS.get(resp_key, resp_key)
            defs  = GOAL_DEFAULTS.get(resp_key, ("target", "", "", ""))

            row_w = QWidget(); row_w.setStyleSheet("background:transparent;")
            row_l = QHBoxLayout(row_w)
            row_l.setContentsMargins(0,2,0,2); row_l.setSpacing(4)

            # İsim
            nm = QLabel(label)
            nm.setFixedWidth(130)
            nm.setStyleSheet(f"color:{TXT}; font-size:11px; background:transparent;")
            row_l.addWidget(nm)

            # Hedef combo
            goal_cb = QComboBox(); goal_cb.setFixedWidth(90); goal_cb.setFixedHeight(26)
            goal_cb.addItems(GOAL_OPTS)
            goal_cb.setCurrentText(defs[0])
            row_l.addWidget(goal_cb)

            # LSL
            lsl_e = QLineEdit(defs[1]); lsl_e.setFixedWidth(72); lsl_e.setFixedHeight(26)
            lsl_e.setPlaceholderText("LSL")
            row_l.addWidget(lsl_e)

            # USL
            usl_e = QLineEdit(defs[2]); usl_e.setFixedWidth(72); usl_e.setFixedHeight(26)
            usl_e.setPlaceholderText("USL")
            row_l.addWidget(usl_e)

            # Hedef değer
            tgt_e = QLineEdit(defs[3]); tgt_e.setFixedWidth(80); tgt_e.setFixedHeight(26)
            tgt_e.setPlaceholderText("Hedef")
            row_l.addWidget(tgt_e)

            # Ağırlık s
            s_sp = QDoubleSpinBox()
            s_sp.setRange(0.1, 5.0); s_sp.setValue(1.0); s_sp.setDecimals(1)
            s_sp.setSingleStep(0.1); s_sp.setFixedWidth(72); s_sp.setFixedHeight(26)
            row_l.addWidget(s_sp)

            self.spec_vlay.insertWidget(self.spec_vlay.count()-1, row_w)
            self._spec_rows[resp_key] = {
                "goal": goal_cb, "lsl": lsl_e,
                "usl": usl_e, "target": tgt_e, "weight": s_sp,
            }

    def _collect_specs(self):
        specs = {}
        for resp_key, widgets in self._spec_rows.items():
            def safe_float(txt):
                try: return float(txt.strip().replace(",", "."))
                except: return None
            specs[resp_key] = {
                "goal":   widgets["goal"].currentText(),
                "lsl":    safe_float(widgets["lsl"].text()),
                "usl":    safe_float(widgets["usl"].text()),
                "target": safe_float(widgets["target"].text()),
                "weight": widgets["weight"].value(),
            }
            self.project.spec_limits[resp_key] = specs[resp_key]
        return specs

    # ── Optimizasyon ─────────────────────────────────────────────────────────
    def _run_optimization(self):
        if not self.project.model_results:
            QMessageBox.warning(self, "", "Önce Model sekmesinden model kurun.")
            return
        if not self._spec_rows:
            QMessageBox.warning(self, "", "Önce bu sekmeye gelip spesifikasyonları doldurun.")
            return

        specs = self._collect_specs()
        self.btn_opt.setEnabled(False)
        self.btn_opt.setText("⏳ Optimize ediliyor...")
        self.status_lbl.setText("Optimizasyon çalışıyor — lütfen bekleyin...")
        self.app.status_bar.showMessage(
            "Diferansiyel evrim optimizasyonu çalışıyor... (30-60 sn sürebilir)")

        self._worker = OptWorker(self.project, specs)
        self._worker.finished.connect(self._on_opt_done)
        self._worker.start()

    def _on_opt_done(self, solutions, err):
        self.btn_opt.setEnabled(True)
        self.btn_opt.setText("▶  Optimize Et")

        if err:
            QMessageBox.critical(self, "Hata", f"Optimizasyon hatası:\n{err}")
            return
        if not solutions:
            n_runs = len(self.project.design_matrix) if self.project.design_matrix is not None else 0
            n_factors = len(self.project.factors)
            min_runs = n_factors + 2
            if n_runs <= n_factors + 1:
                msg = (
                    "Optimum cozum bulunamadi.\n\n"
                    "Sebep: " + str(n_runs) + " run ile " + str(n_factors) +
                    " faktorlu model icin yeterli veri yok"
                    " (en az " + str(min_runs) + " run gerekli).\n\n"
                    "Oneri: CCD veya Box-Behnken tasarimi kullanin.")
            else:
                msg = (
                    "Optimum cozum bulunamadi.\n\n"
                    "Olasi sebepler:\n"
                    "- Model R2 cok dusuk\n"
                    "- Spesifikasyon sinirlar cok dar\n"
                    "- Faktor araliklari yetersiz\n\n"
                    "Oneri: USL/LSL sinirlarini genisletin.")
            QMessageBox.warning(self, "Cozum Bulunamadi", msg)
            self.status_lbl.setText("Cozum bulunamadi - spesifikasyonlari gozden gecirin.")
            return

        self.project.opt_solutions = solutions
        self._show_solutions(solutions)
        self._show_robustness(solutions[0])
        self.status_lbl.setText(
            f"✅  {len(solutions)} çözüm bulundu  |  "
            f"En iyi desirability: {solutions[0]['desirability']:.4f}")
        self.app.status_bar.showMessage(
            f"Optimizasyon tamamlandı — En iyi D={solutions[0]['desirability']:.4f}", 6000)

    # ── Sonuç gösterimi ───────────────────────────────────────────────────────
    def _show_solutions(self, solutions):
        self.result_tabs.clear()

        for rank, sol in enumerate(solutions):
            d = sol["desirability"]
            tab_w = QWidget()
            tab_l = QVBoxLayout(tab_w)
            tab_l.setContentsMargins(10, 10, 10, 10)
            tab_l.setSpacing(8)

            # Desirability skoru
            d_color = GREEN if d >= 0.8 else GOLD if d >= 0.5 else RED
            d_lbl = QLabel(f"Desirability: {d:.4f}")
            d_lbl.setStyleSheet(
                f"color:{d_color}; font-size:16px; font-weight:bold; background:transparent;")
            tab_l.addWidget(d_lbl)

            # Faktör değerleri
            tab_l.addWidget(section_label("Faktör Değerleri"))
            f_tbl = QTableWidget()
            f_tbl.setColumnCount(3)
            f_tbl.setHorizontalHeaderLabels(["Faktör", "Değer", "Birim"])
            f_tbl.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.Stretch)
            f_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            f_tbl.setFixedHeight(min(200, 36 + 28*len(self.project.factors)))
            f_tbl.setRowCount(len(self.project.factors))

            for fi, f in enumerate(self.project.factors):
                val = sol["factors"].get(f["name"], 0)
                for ci, txt in enumerate([f["name"],
                                           f"{val:.4f}",
                                           f.get("unit","")]):
                    item = QTableWidgetItem(txt)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(QColor(BG3))
                    f_tbl.setItem(fi, ci, item)
            tab_l.addWidget(f_tbl)

            # Tahmin değerleri
            tab_l.addWidget(section_label("Tahmin Değerleri (CQA)"))
            p_tbl = QTableWidget()
            p_tbl.setColumnCount(4)
            p_tbl.setHorizontalHeaderLabels(["Yanıt", "Tahmin", "LSL", "USL"])
            p_tbl.horizontalHeader().setSectionResizeMode(
                QHeaderView.ResizeMode.Stretch)
            p_tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            preds = sol.get("predictions", {})
            active = [r for r in self.project.responses
                      if r in self.project.model_results]
            p_tbl.setRowCount(len(active))
            p_tbl.setFixedHeight(min(200, 36 + 28*len(active)))

            for pi, resp_key in enumerate(active):
                pred_val = preds.get(resp_key)
                spec     = self.project.spec_limits.get(resp_key, {})
                lsl = spec.get("lsl"); usl = spec.get("usl")

                # Spec kontrolü
                in_spec = True
                if pred_val is not None:
                    if lsl is not None and pred_val < lsl: in_spec = False
                    if usl is not None and pred_val > usl: in_spec = False

                bg = QColor("#0a2a0a") if in_spec else QColor("#2a0a0a")
                fg = QColor(GREEN)     if in_spec else QColor(RED)

                items = [
                    QTableWidgetItem(RESPONSE_LABELS.get(resp_key, resp_key)),
                    QTableWidgetItem(f"{pred_val:.4f}" if pred_val is not None else "—"),
                    QTableWidgetItem(f"{lsl:.4f}" if lsl is not None else "—"),
                    QTableWidgetItem(f"{usl:.4f}" if usl is not None else "—"),
                ]
                for ci, item in enumerate(items):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(bg)
                    if ci == 1:
                        item.setForeground(fg)
                    p_tbl.setItem(pi, ci, item)
            tab_l.addWidget(p_tbl)
            tab_l.addStretch()

            d_str = f"{d:.3f}"
            self.result_tabs.addTab(tab_w, f"#{rank+1}  D={d_str}")

        # Desirability bar grafiği
        self._plot_desirability(solutions)

    def _plot_desirability(self, solutions):
        self.fig_d.clear()
        self.fig_d.patch.set_facecolor(BG2)
        ax = self.fig_d.add_subplot(111)
        ax.set_facecolor(BG3)
        ax.tick_params(colors=TXT2, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor("#2a4060")

        labels = [f"#{i+1}" for i in range(len(solutions))]
        values = [s["desirability"] for s in solutions]
        colors = [GREEN if v >= 0.8 else GOLD if v >= 0.5 else RED for v in values]

        bars = ax.bar(labels, values, color=colors, alpha=0.85,
                      edgecolor="#2a4060", linewidth=0.8)
        ax.set_ylim(0, 1.05)
        ax.axhline(0.8, color=GREEN,  lw=1, linestyle="--", alpha=0.6)
        ax.axhline(0.5, color=GOLD,   lw=1, linestyle="--", alpha=0.6)
        ax.set_ylabel("Desirability", color=TXT2, fontsize=8)
        ax.set_title("Top-5 Desirability Skoru", color=GOLD, fontsize=9)

        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width()/2, val + 0.02,
                    f"{val:.3f}", ha="center", va="bottom",
                    color=TXT, fontsize=8)

        self.fig_d.tight_layout(pad=1.0)
        self.canvas_d.draw()

    def _show_robustness(self, best_sol):
        """Optimum ±10% sapma senaryosu."""
        import statsmodels.api as sm
        p      = self.project
        safe   = p.get_safe_names()
        factors = p.factors

        def encode(val, f):
            lo, hi = f["low"], f["high"]
            mid = f.get("mid", (lo+hi)/2)
            if hi == lo: return 0.0
            if val <= mid and (mid-lo) != 0:
                return (val-lo)/(mid-lo) - 1
            elif (hi-mid) != 0:
                return (val-mid)/(hi-mid)
            return 0.0

        def predict_at(resp_key, x_vals):
            res = p.model_results.get(resp_key)
            if res is None: return None
            model = res["model"]
            row = {}
            for i, f in enumerate(factors):
                row[safe[i]] = encode(x_vals[i], f)
            for name in safe:
                row[f"{name}_sq"] = row[name]**2
            for a in range(len(factors)):
                for b in range(a+1, len(factors)):
                    row[f"{safe[a]}_x_{safe[b]}"] = row[safe[a]]*row[safe[b]]
            df_row = pd.DataFrame([row])
            row_mat = sm.add_constant(df_row, has_constant="add")
            for col in model.params.index:
                if col not in row_mat.columns:
                    row_mat[col] = 0.0
            row_mat = row_mat.reindex(columns=model.params.index, fill_value=0.0)
            try:
                return float(model.predict(row_mat)[0])
            except Exception:
                return None

        x_opt  = best_sol["x"]
        active = [r for r in p.responses if r in p.model_results]
        self.robust_table.setRowCount(len(active))

        for ri, resp_key in enumerate(active):
            label = RESPONSE_LABELS.get(resp_key, resp_key)
            opt_pred = best_sol["predictions"].get(resp_key)

            # ±10% senaryolar
            preds_low, preds_high = [], []
            for i, f in enumerate(factors):
                span = f["high"] - f["low"]
                step = span * 0.10

                x_lo = list(x_opt); x_lo[i] = max(f["low"], x_opt[i] - step)
                x_hi = list(x_opt); x_hi[i] = min(f["high"], x_opt[i] + step)
                pv_lo = predict_at(resp_key, x_lo)
                pv_hi = predict_at(resp_key, x_hi)
                if pv_lo is not None: preds_low.append(pv_lo)
                if pv_hi is not None: preds_high.append(pv_hi)

            lo_val = min(preds_low)  if preds_low  else None
            hi_val = max(preds_high) if preds_high else None

            spec = p.spec_limits.get(resp_key, {})
            lsl  = spec.get("lsl"); usl = spec.get("usl")

            def in_spec_check(v):
                if v is None: return True
                if lsl is not None and v < lsl: return False
                if usl is not None and v > usl: return False
                return True

            items = [
                QTableWidgetItem(label),
                QTableWidgetItem(f"{opt_pred:.4f}" if opt_pred is not None else "—"),
                QTableWidgetItem(f"{lo_val:.4f}"   if lo_val  is not None else "—"),
                QTableWidgetItem(f"{hi_val:.4f}"   if hi_val  is not None else "—"),
            ]
            for ci, item in enumerate(items):
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                ok = in_spec_check(lo_val) and in_spec_check(hi_val)
                item.setBackground(QColor("#0a2a0a" if ok else "#2a1a0a"))
                if ci in (2, 3):
                    item.setForeground(QColor(GREEN if ok else GOLD))
                self.robust_table.setItem(ri, ci, item)


# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 6 — DESIGN SPACE (ICH Q8)
# ═══════════════════════════════════════════════════════════════════════════════

class DesignSpaceWorker(QThread):
    """Design Space ızgara hesabını arka planda çalıştırır."""
    finished = pyqtSignal(object, object, dict, int, str)  # XX, YY, grids, n, mode
    error    = pyqtSignal(str)

    def __init__(self, tab, xi, yi, n, mode):
        super().__init__()
        self.tab  = tab
        self.xi   = xi
        self.yi   = yi
        self.n    = n
        self.mode = mode

    def run(self):
        try:
            p      = self.tab.project
            active = [r for r in p.responses
                      if r in p.model_results and r in p.spec_limits]
            if not active:
                self.error.emit("Hiçbir yanıt için model + spesifikasyon bulunamadı.")
                return

            grids = {}
            for resp_key in active:
                XX, YY, ZZ = self.tab._predict_grid(
                    resp_key, self.xi, self.yi, self.n)
                grids[resp_key] = ZZ

            self.finished.emit(XX, YY, grids, self.n, self.mode)
        except Exception as e:
            import traceback
            write_log(f"DesignSpaceWorker hata:\n{traceback.format_exc()}")
            self.error.emit(str(e))

class Tab6_DesignSpace(QWidget):
    """ICH Q8 uyumlu Design Space haritası — tüm CQA spec içi bölge."""

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project    = project
        self.app        = app_ref
        self._ds_worker = None
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(10)

        # ── Kontrol çubuğu ────────────────────────────────────────────────────
        ctrl = card_frame()
        cl = QHBoxLayout(ctrl)
        cl.setContentsMargins(14, 8, 14, 8)
        cl.setSpacing(14)

        cl.addWidget(QLabel("X Ekseni:"))
        self.x_combo = QComboBox(); self.x_combo.setFixedHeight(28)
        cl.addWidget(self.x_combo)

        cl.addWidget(QLabel("Y Ekseni:"))
        self.y_combo = QComboBox(); self.y_combo.setFixedHeight(28)
        cl.addWidget(self.y_combo)

        cl.addWidget(QLabel("Çözünürlük:"))
        self.res_spin = QSpinBox()
        self.res_spin.setRange(20, 120); self.res_spin.setValue(60)
        self.res_spin.setFixedWidth(65); self.res_spin.setFixedHeight(28)
        cl.addWidget(self.res_spin)

        # Görüntüleme modu
        cl.addWidget(QLabel("Mod:"))
        self.mode_combo = QComboBox(); self.mode_combo.setFixedHeight(28)
        self.mode_combo.addItems([
            "Tüm CQA (AND)",
            "CQA sayısı (heatmap)",
            "Tek CQA seç",
        ])
        self.mode_combo.currentTextChanged.connect(self._on_mode_changed)
        cl.addWidget(self.mode_combo)

        self.single_resp_combo = QComboBox()
        self.single_resp_combo.setFixedHeight(28)
        self.single_resp_combo.setVisible(False)
        cl.addWidget(self.single_resp_combo)

        cl.addStretch()

        self.btn_plot = make_btn("▶  Harita Çiz", "rgba(20,80,20,0.9)", 32)
        self.btn_plot.setStyleSheet(self.btn_plot.styleSheet() +
                                    "font-size:12px; font-weight:bold;")
        self.btn_plot.clicked.connect(self._plot)
        cl.addWidget(self.btn_plot)
        outer.addWidget(ctrl)

        # ── Sabit faktör slider'ları ───────────────────────────────────────────
        self.slider_card = card_frame()
        self.slider_layout = QHBoxLayout(self.slider_card)
        self.slider_layout.setContentsMargins(14, 8, 14, 8)
        self.slider_layout.setSpacing(20)
        self.slider_card.setFixedHeight(70)
        self._slider_widgets = {}
        outer.addWidget(self.slider_card)

        # ── Ana grafik (sekmeli) ──────────────────────────────────────────────
        self.graph_tabs = QTabWidget()
        self.graph_tabs.setStyleSheet(
            f"QTabBar::tab {{ padding:4px 12px; font-size:11px; }}")

        # Sekme 1: Design Space haritası
        map_w = QWidget()
        map_l = QVBoxLayout(map_w)
        map_l.setContentsMargins(0,0,0,0)
        self.fig = Figure(figsize=(14, 6), facecolor=BG2)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setStyleSheet(f"background:{BG2};")
        map_l.addWidget(self.canvas)
        self.graph_tabs.addTab(map_w, "🗺  Design Space Haritası")

        # Sekme 2: 3D Faktör Uzayı (3+ faktörde dolu olacak)
        self.scatter3d_w = QWidget()
        s3d_l = QVBoxLayout(self.scatter3d_w)
        s3d_l.setContentsMargins(0,0,0,0)
        self.fig3d = Figure(figsize=(14, 6), facecolor=BG2)
        self.canvas3d = FigureCanvas(self.fig3d)
        self.canvas3d.setStyleSheet(f"background:{BG2};")
        s3d_l.addWidget(self.canvas3d)
        # Sıfırla butonu
        s3d_bot = QHBoxLayout()
        s3d_bot.addStretch()
        btn_reset3d = make_btn("↺  Görünümü Sıfırla", "rgba(30,50,90,0.8)", 28)
        btn_reset3d.setToolTip("3D grafiği başlangıç açısına döndür")
        btn_reset3d.clicked.connect(self._reset_3d_scatter_view)
        s3d_bot.addWidget(btn_reset3d)
        s3d_bot.setContentsMargins(0,4,8,4)
        s3d_bot_w = QWidget(); s3d_bot_w.setLayout(s3d_bot)
        s3d_l.addWidget(s3d_bot_w)
        self.graph_tabs.addTab(self.scatter3d_w, "🔵  3D Faktör Uzayı")

        outer.addWidget(self.graph_tabs, 1)

        # ── Alt çubuk ────────────────────────────────────────────────────────
        bot = QHBoxLayout()
        self.status_lbl = QLabel(
            "Model ve spesifikasyon sınırları tanımlandıktan sonra harita çizebilirsiniz.")
        self.status_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        bot.addWidget(self.status_lbl)
        bot.addStretch()
        self.btn_next = make_btn("Doğrulama  ▶", "rgba(20,80,20,0.8)", 34)
        self.btn_next.setStyleSheet(self.btn_next.styleSheet() +
                                    "font-size:12px; font-weight:bold;")
        self.btn_next.clicked.connect(lambda: self.app.tabs.setCurrentIndex(6))
        bot.addWidget(self.btn_next)
        outer.addLayout(bot)

        self.x_combo.currentTextChanged.connect(self._on_axis_changed)
        self.y_combo.currentTextChanged.connect(self._on_axis_changed)

    # ── Public ───────────────────────────────────────────────────────────────
    def refresh(self):
        if not self.project.model_results:
            return
        self._populate_combos()

    # ── İç metodlar ──────────────────────────────────────────────────────────
    def _populate_combos(self):
        self.x_combo.blockSignals(True)
        self.y_combo.blockSignals(True)
        self.single_resp_combo.blockSignals(True)

        self.x_combo.clear(); self.y_combo.clear()
        for i, f in enumerate(self.project.factors):
            label = f"{f['name']}" + (f" ({f['unit']})" if f.get("unit") else "")
            self.x_combo.addItem(label, i)
            self.y_combo.addItem(label, i)
        if self.y_combo.count() >= 2:
            self.y_combo.setCurrentIndex(1)

        self.single_resp_combo.clear()
        for r in self.project.responses:
            if r in self.project.model_results:
                self.single_resp_combo.addItem(RESPONSE_LABELS.get(r, r), r)

        self.x_combo.blockSignals(False)
        self.y_combo.blockSignals(False)
        self.single_resp_combo.blockSignals(False)
        self._build_sliders()

    def _on_mode_changed(self, mode):
        self.single_resp_combo.setVisible("Tek CQA" in mode)

    def _on_axis_changed(self):
        self._build_sliders()

    def _build_sliders(self):
        while self.slider_layout.count():
            item = self.slider_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._slider_widgets = {}

        xi = self.x_combo.currentData() or 0
        yi = self.y_combo.currentData() or 1

        for i, f in enumerate(self.project.factors):
            if i in (xi, yi):
                continue
            mid = f.get("mid", (f["low"] + f["high"]) / 2)
            grp = QGroupBox(f["name"])
            grp.setFixedWidth(150)
            gl = QVBoxLayout(grp)
            gl.setContentsMargins(6, 4, 6, 4)
            sp = QDoubleSpinBox()
            sp.setRange(f["low"], f["high"])
            sp.setValue(mid)
            sp.setDecimals(3)
            sp.setSingleStep((f["high"] - f["low"]) / 20)
            sp.setFixedHeight(26)
            if f.get("unit"):
                sp.setSuffix(f"  {f['unit']}")
            gl.addWidget(sp)
            self.slider_layout.addWidget(grp)
            self._slider_widgets[i] = sp
        self.slider_layout.addStretch()

    def _get_fixed(self):
        return {i: sp.value() for i, sp in self._slider_widgets.items()}

    def _encode(self, val, f):
        lo, hi = f["low"], f["high"]
        mid = f.get("mid", (lo + hi) / 2)
        if hi == lo: return 0.0
        if val <= mid and (mid - lo) != 0:
            return (val - lo) / (mid - lo) - 1
        elif (hi - mid) != 0:
            return (val - mid) / (hi - mid)
        return 0.0

    def _predict_grid(self, resp_key, xi, yi, n):
        import statsmodels.api as _sm
        p = self.project
        safe = p.get_safe_names()
        factors = p.factors
        n_factors = len(factors)
        res = p.model_results[resp_key]
        model = res["model"]

        fx = factors[xi]; fy = factors[yi]
        x_arr = np.linspace(fx["low"], fx["high"], n)
        y_arr = np.linspace(fy["low"], fy["high"], n)
        XX, YY = np.meshgrid(x_arr, y_arr)
        ZZ = np.full_like(XX, np.nan)

        fixed = self._get_fixed()
        param_idx = model.params.index

        for r in range(n):
            for c in range(n):
                row = {}
                for k, f in enumerate(factors):
                    if k == xi:
                        row[safe[k]] = self._encode(XX[r, c], f)
                    elif k == yi:
                        row[safe[k]] = self._encode(YY[r, c], f)
                    else:
                        row[safe[k]] = self._encode(
                            fixed.get(k, f.get("mid", (f["low"]+f["high"])/2)), f)
                for name in safe:
                    row[f"{name}_sq"] = row[name] ** 2
                for a in range(n_factors):
                    for b in range(a+1, n_factors):
                        row[f"{safe[a]}_x_{safe[b]}"] = row[safe[a]] * row[safe[b]]

                df_row = pd.DataFrame([row])
                row_mat = _sm.add_constant(df_row, has_constant="add")
                for col in param_idx:
                    if col not in row_mat.columns:
                        row_mat[col] = 0.0
                row_mat = row_mat.reindex(columns=param_idx, fill_value=0.0)
                try:
                    ZZ[r, c] = float(model.predict(row_mat)[0])
                except Exception:
                    pass

        return XX, YY, ZZ

    def _check_spec(self, val, spec):
        """Tek bir değerin spec içinde olup olmadığını kontrol et."""
        if val is None or np.isnan(val):
            return False
        lsl = spec.get("lsl"); usl = spec.get("usl")
        if lsl is not None and val < lsl: return False
        if usl is not None and val > usl: return False
        return True

    def _plot(self):
        if not self.project.model_results:
            QMessageBox.warning(self, "", "Önce Model sekmesinden model kurun.")
            return
        if not self.project.spec_limits:
            QMessageBox.warning(self, "",
                "Önce Optimizasyon sekmesinden spesifikasyon sınırlarını tanımlayın.")
            return

        xi = self.x_combo.currentData()
        yi = self.y_combo.currentData()
        if xi is None or yi is None or xi == yi:
            QMessageBox.warning(self, "", "X ve Y farklı faktörler olmalıdır.")
            return

        self.btn_plot.setEnabled(False)
        self.btn_plot.setText("⏳ Hesaplanıyor...")
        self.app.status_bar.showMessage(
            "Design Space hesaplanıyor — lütfen bekleyin...")

        n    = self.res_spin.value()
        mode = self.mode_combo.currentText()

        self._ds_worker = DesignSpaceWorker(self, xi, yi, n, mode)
        self._ds_worker.finished.connect(self._on_ds_done)
        self._ds_worker.error.connect(self._on_ds_error)
        self._ds_worker.start()

    def _compute_and_draw(self, xi, yi, n, mode):
        p = self.project
        factors = p.factors
        fx = factors[xi]; fy = factors[yi]
        specs = p.spec_limits

        # Aktif yanıtlar (hem model hem spec tanımlı)
        active = [r for r in p.responses
                  if r in p.model_results and r in specs]

        if not active:
            QMessageBox.warning(self, "",
                "Hiçbir yanıt için model + spesifikasyon bulunamadı.")
            return

        # Her aktif yanıt için grid tahmin
        grids = {}
        for resp_key in active:
            self.app.status_bar.showMessage(
                f"Hesaplanıyor: {RESPONSE_LABELS.get(resp_key, resp_key)}...")
            QApplication.processEvents()
            XX, YY, ZZ = self._predict_grid(resp_key, xi, yi, n)
            grids[resp_key] = ZZ

        self.fig.clear()
        self.fig.patch.set_facecolor(BG2)

        if mode == "Tek CQA seç":
            self._draw_single(XX, YY, grids, xi, yi, active)
        elif mode == "CQA sayısı (heatmap)":
            self._draw_count(XX, YY, grids, active, xi, yi)
        else:  # Tüm CQA (AND)
            self._draw_and(XX, YY, grids, active, xi, yi)

        self.fig.tight_layout(pad=2.0)
        self.canvas.draw()

        # Durum bilgisi
        if mode == "Tüm CQA (AND)":
            and_mask = np.ones_like(XX, dtype=bool)
            for resp_key, ZZ in grids.items():
                spec = specs[resp_key]
                for r in range(n):
                    for c in range(n):
                        if not self._check_spec(ZZ[r, c], spec):
                            and_mask[r, c] = False
            pct = 100 * and_mask.sum() / and_mask.size
            self.status_lbl.setText(
                f"✅  Design Space: Alan %{pct:.1f} spec içinde  |  "
                f"{len(active)} CQA  |  "
                f"X: {fx['name']}  Y: {fy['name']}")
        else:
            self.status_lbl.setText(
                f"✅  Harita hazır  |  {len(active)} CQA  |  "
                f"X: {fx['name']}  Y: {fy['name']}")

    def _draw_and(self, XX, YY, grids, active, xi, yi):
        """Tüm CQA'ların aynı anda spec içinde olduğu bölge — ICH Q8."""
        p = self.project
        specs = p.spec_limits
        n = XX.shape[0]

        # Yeşil=tümü OK, Sarı=1 hariç, Kırmızı=2+ hariç
        fail_count = np.zeros_like(XX, dtype=int)
        for resp_key, ZZ in grids.items():
            spec = specs[resp_key]
            for r in range(n):
                for c in range(n):
                    if not self._check_spec(ZZ[r, c], spec):
                        fail_count[r, c] += 1

        # Sol: Design Space haritası
        ax = self.fig.add_subplot(1, 2, 1)
        ax.set_facecolor(BG3)
        ax.tick_params(colors=TXT2, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor("#2a4060")

        import matplotlib.colors as mcolors
        cmap = mcolors.ListedColormap([
            "#003300",   # 0 fail = koyu yeşil
            "#1a5200",   # 1 fail = açık yeşil
            "#7a5200",   # 2 fail = sarı-kahve
            "#8b0000",   # 3+ fail = kırmızı
        ])
        bounds_c = [-0.5, 0.5, 1.5, 2.5, max(len(active)+0.5, 3.5)]
        norm = mcolors.BoundaryNorm(bounds_c, cmap.N)

        im = ax.pcolormesh(XX, YY, fail_count, cmap=cmap, norm=norm,
                           shading="auto", alpha=0.9)

        # Design Space sınırı (0 fail bölgesi kontur çizgisi)
        try:
            ds_mask = (fail_count == 0).astype(float)
            ax.contour(XX, YY, ds_mask, levels=[0.5],
                       colors=[GOLD], linewidths=2.0)
        except Exception:
            pass

        # Optimum nokta
        self._add_optimum(ax, xi, yi)

        # Gerçek run noktaları
        self._add_run_points(ax, xi, yi)

        fx = self.project.factors[xi]; fy = self.project.factors[yi]
        ax.set_xlabel(f"{fx['name']}" + (f" ({fx['unit']})" if fx.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_ylabel(f"{fy['name']}" + (f" ({fy['unit']})" if fy.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_title("Design Space  (ICH Q8)", color=GOLD, fontsize=10)

        cb = self.fig.colorbar(im, ax=ax, ticks=[0, 1, 2, 3])
        cb.ax.set_yticklabels(["Tümü OK", "1 Hata", "2 Hata", "3+"], fontsize=7)
        cb.ax.tick_params(colors=TXT2)

        # Sağ: CQA uygunluk özet çubuğu
        ax2 = self.fig.add_subplot(1, 2, 2)
        ax2.set_facecolor(BG3)
        ax2.tick_params(colors=TXT2, labelsize=8)
        for spine in ax2.spines.values():
            spine.set_edgecolor("#2a4060")

        labels, ok_pcts = [], []
        for resp_key, ZZ in grids.items():
            spec = self.project.spec_limits[resp_key]
            ok_count = sum(
                1 for r in range(n) for c in range(n)
                if self._check_spec(ZZ[r, c], spec))
            ok_pcts.append(100 * ok_count / (n*n))
            labels.append(RESPONSE_LABELS.get(resp_key, resp_key))

        colors = [GREEN if v >= 80 else GOLD if v >= 50 else RED for v in ok_pcts]
        bars = ax2.barh(labels, ok_pcts, color=colors, alpha=0.85,
                        edgecolor="#2a4060", linewidth=0.8)
        ax2.set_xlim(0, 105)
        ax2.axvline(80, color=GOLD, lw=1, linestyle="--", alpha=0.6)
        ax2.set_xlabel("Spec İçi Alan (%)", color=TXT2, fontsize=8)
        ax2.set_title("CQA Uygunluk Oranı", color=GOLD, fontsize=10)
        for bar, val in zip(bars, ok_pcts):
            ax2.text(val + 1, bar.get_y() + bar.get_height()/2,
                     f"{val:.1f}%", va="center", color=TXT, fontsize=8)

    def _draw_count(self, XX, YY, grids, active, xi, yi):
        """Kaç CQA spec içinde — ısı haritası."""
        p = self.project; specs = p.spec_limits; n = XX.shape[0]
        ok_count = np.zeros_like(XX, dtype=float)
        for resp_key, ZZ in grids.items():
            spec = specs[resp_key]
            for r in range(n):
                for c in range(n):
                    if self._check_spec(ZZ[r, c], spec):
                        ok_count[r, c] += 1

        ax = self.fig.add_subplot(1, 1, 1)
        ax.set_facecolor(BG3)
        ax.tick_params(colors=TXT2, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor("#2a4060")

        cf = ax.contourf(XX, YY, ok_count,
                         levels=np.arange(-0.5, len(active)+1.5, 1),
                         cmap="RdYlGn", alpha=0.9)
        cb = self.fig.colorbar(cf, ax=ax)
        cb.ax.tick_params(colors=TXT2, labelsize=7)
        cb.set_label("Spec İçi CQA Sayısı", color=TXT2, fontsize=8)

        self._add_optimum(ax, xi, yi)
        self._add_run_points(ax, xi, yi)

        fx = self.project.factors[xi]; fy = self.project.factors[yi]
        ax.set_xlabel(f"{fx['name']}" + (f" ({fx['unit']})" if fx.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_ylabel(f"{fy['name']}" + (f" ({fy['unit']})" if fy.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_title(f"Spec İçi CQA Sayısı (maks={len(active)})",
                     color=GOLD, fontsize=10)

    def _draw_single(self, XX, YY, grids, xi, yi, active):
        """Tek CQA için yüzey + spec sınırı."""
        resp_key = self.single_resp_combo.currentData()
        if resp_key is None or resp_key not in grids:
            resp_key = active[0]

        ZZ   = grids[resp_key]
        spec = self.project.spec_limits.get(resp_key, {})
        lsl  = spec.get("lsl"); usl = spec.get("usl")
        label = RESPONSE_LABELS.get(resp_key, resp_key)

        ax = self.fig.add_subplot(1, 1, 1)
        ax.set_facecolor(BG3)
        ax.tick_params(colors=TXT2, labelsize=8)
        for spine in ax.spines.values():
            spine.set_edgecolor("#2a4060")

        cf = ax.contourf(XX, YY, ZZ, levels=25, cmap="viridis", alpha=0.9)
        cs = ax.contour(XX, YY, ZZ, levels=25, colors="white",
                        alpha=0.2, linewidths=0.4)

        # Spec sınır konturları
        if lsl is not None:
            try:
                ax.contour(XX, YY, ZZ, levels=[lsl],
                           colors=["#ff4444"], linewidths=2.0,
                           linestyles="--")
                ax.clabel(ax.contour(XX, YY, ZZ, levels=[lsl],
                                     colors=["#ff4444"]),
                          fmt=f"LSL={lsl}", fontsize=7, colors="#ff4444")
            except Exception:
                pass
        if usl is not None:
            try:
                ax.contour(XX, YY, ZZ, levels=[usl],
                           colors=["#ff8800"], linewidths=2.0,
                           linestyles="--")
            except Exception:
                pass

        cb = self.fig.colorbar(cf, ax=ax)
        cb.ax.tick_params(colors=TXT2, labelsize=7)
        cb.set_label(label, color=TXT2, fontsize=8)

        self._add_optimum(ax, xi, yi)
        self._add_run_points(ax, xi, yi)

        fx = self.project.factors[xi]; fy = self.project.factors[yi]
        ax.set_xlabel(f"{fx['name']}" + (f" ({fx['unit']})" if fx.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_ylabel(f"{fy['name']}" + (f" ({fy['unit']})" if fy.get("unit") else ""),
                      color=TXT2, fontsize=9)
        ax.set_title(f"{label} — Design Space", color=GOLD, fontsize=10)

    def _add_optimum(self, ax, xi, yi):
        """Optimum noktayı grafik üzerine ekle."""
        if not self.project.opt_solutions:
            return
        best = self.project.opt_solutions[0]
        fx = self.project.factors[xi]; fy = self.project.factors[yi]
        ox = best.get("factors", {}).get(fx["name"])
        oy = best.get("factors", {}).get(fy["name"])
        if ox is not None and oy is not None:
            ax.plot(ox, oy, "*", color=GOLD, markersize=16,
                    zorder=5, label=f"Optimum (D={best['desirability']:.3f})",
                    markeredgecolor="white", markeredgewidth=0.8)
            ax.legend(fontsize=8, labelcolor=TXT2,
                      facecolor=BG2, edgecolor="#2a4060",
                      loc="upper right")

    def _add_run_points(self, ax, xi, yi):
        """Gerçek run noktalarını ekle."""
        if self.project.design_matrix is None:
            return
        df = self.project.design_matrix
        fx = self.project.factors[xi]; fy = self.project.factors[yi]
        xs, ys = [], []
        for ri in range(len(df)):
            if self.project.run_results.get(ri):
                xs.append(df.iloc[ri][fx["name"]])
                ys.append(df.iloc[ri][fy["name"]])
        if xs:
            ax.scatter(xs, ys, c="white", s=25, zorder=4,
                       edgecolors="#4a7ab0", linewidths=0.8,
                       alpha=0.85, label="Ölçüm noktaları")

    def _on_ds_done(self, XX, YY, grids, n, mode):
        """Worker tamamlandığında UI'ı güncelle."""
        try:
            self.fig.clear()
            self.fig.patch.set_facecolor(BG2)
            xi = self.x_combo.currentData()
            yi = self.y_combo.currentData()

            if mode == "Tek CQA seç":
                self._draw_single(XX, YY, grids, xi, yi,
                                  list(grids.keys()))
            elif mode == "CQA sayısı (heatmap)":
                self._draw_count(XX, YY, grids,
                                 list(grids.keys()), xi, yi)
            else:
                self._draw_and(XX, YY, grids,
                               list(grids.keys()), xi, yi)

            self.fig.tight_layout(pad=2.0)
            self.canvas.draw()

            # Durum bilgisi
            fx = self.project.factors[xi]
            fy = self.project.factors[yi]
            if mode == "Tüm CQA (AND)":
                specs = self.project.spec_limits
                fail_count = np.zeros((n, n), dtype=int)
                for resp_key, ZZ in grids.items():
                    spec = specs.get(resp_key, {})
                    for r in range(n):
                        for c in range(n):
                            if not self._check_spec(ZZ[r, c], spec):
                                fail_count[r, c] += 1
                pct = 100 * (fail_count == 0).sum() / fail_count.size
                self.status_lbl.setText(
                    f"✅  Design Space: Alan %{pct:.1f} spec içinde  |  "
                    f"{len(grids)} CQA  |  X: {fx['name']}  Y: {fy['name']}")
            else:
                self.status_lbl.setText(
                    f"✅  Harita hazır  |  {len(grids)} CQA  |  "
                    f"X: {fx['name']}  Y: {fy['name']}")

            self.app.status_bar.showMessage("Design Space hazır.", 4000)

            # 3D Scatter — 3+ faktörde çiz
            if len(self.project.factors) >= 3:
                self._draw_3d_scatter(grids, xi, yi)
        except Exception as e:
            import traceback
            write_log(f"_on_ds_done hata:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Hata", f"Grafik çizilemedi:\n{e}")
        finally:
            self.btn_plot.setEnabled(True)
            self.btn_plot.setText("▶  Harita Çiz")

    def _on_ds_error(self, msg):
        """Worker hata verdiğinde."""
        self.btn_plot.setEnabled(True)
        self.btn_plot.setText("▶  Harita Çiz")
        QMessageBox.critical(self, "Hata", f"Design Space hesaplanamadı:\n{msg}")
        self.app.status_bar.showMessage("Hata oluştu.", 4000)

    def _reset_3d_scatter_view(self):
        """3D scatter grafiğini varsayılan açıya döndür."""
        try:
            for ax in self.fig3d.get_axes():
                if hasattr(ax, 'elev'):
                    ax.view_init(elev=25, azim=-60)
            self.canvas3d.draw()
        except Exception:
            pass

    def _draw_3d_scatter(self, grids, xi, yi):
        """3D faktör uzayı scatter — run noktaları yeşil/kırmızı/sarı."""
        from mpl_toolkits.mplot3d import Axes3D
        p = self.project
        factors = p.factors
        specs   = p.spec_limits
        n_f     = len(factors)
        dm      = p.design_matrix

        # 3'ten az faktör varsa bilgi mesajı göster
        if n_f < 3:
            self.fig3d.clear()
            self.fig3d.patch.set_facecolor(BG2)
            ax = self.fig3d.add_subplot(111)
            ax.set_facecolor(BG2)
            ax.axis("off")
            ax.text(0.5, 0.55, "3D Faktör Uzayı",
                    ha="center", va="center", fontsize=16,
                    color=GOLD, fontweight="bold",
                    transform=ax.transAxes)
            ax.text(0.5, 0.45,
                    "Bu gorunum en az 3 faktor gerektirir.\n"
                    "Mevcut faktor sayisi: " + str(n_f),
                    ha="center", va="center", fontsize=11,
                    color=TXT2, transform=ax.transAxes)
            self.canvas3d.draw()
            return

        if dm is None:
            return

        self.fig3d.clear()
        self.fig3d.patch.set_facecolor(BG2)
        ax = self.fig3d.add_subplot(111, projection='3d')
        ax.set_facecolor(BG3)
        ax.tick_params(colors=TXT2, labelsize=7)
        ax.xaxis.pane.fill = False
        ax.yaxis.pane.fill = False
        ax.zaxis.pane.fill = False
        ax.xaxis.pane.set_edgecolor("#2a4060")
        ax.yaxis.pane.set_edgecolor("#2a4060")
        ax.zaxis.pane.set_edgecolor("#2a4060")

        # 3. faktör indeksi (xi ve yi dışında ilk faktör)
        zi = next((i for i in range(n_f) if i not in (xi, yi)), 0)

        fx = factors[xi]; fy = factors[yi]; fz = factors[zi]
        xs_ok, ys_ok, zs_ok = [], [], []
        xs_fail, ys_fail, zs_fail = [], [], []
        xs_miss, ys_miss, zs_miss = [], [], []

        for ri in range(len(dm)):
            xv = float(dm.iloc[ri][fx["name"]])
            yv = float(dm.iloc[ri][fy["name"]])
            zv = float(dm.iloc[ri][fz["name"]])
            run_data = p.run_results.get(ri, {})

            if not run_data:
                xs_miss.append(xv); ys_miss.append(yv); zs_miss.append(zv)
                continue

            all_ok = True
            for resp_key, spec in specs.items():
                pred_val = run_data.get(resp_key)
                if pred_val is not None:
                    if not self._check_spec(float(pred_val), spec):
                        all_ok = False
                        break

            if all_ok:
                xs_ok.append(xv); ys_ok.append(yv); zs_ok.append(zv)
            else:
                xs_fail.append(xv); ys_fail.append(yv); zs_fail.append(zv)

        # Çiz
        if xs_ok:
            ax.scatter(xs_ok, ys_ok, zs_ok, c=GREEN, s=60,
                       alpha=0.9, edgecolors="white", linewidths=0.5,
                       label="Spec içi ✅", zorder=5)
        if xs_fail:
            ax.scatter(xs_fail, ys_fail, zs_fail, c=RED, s=60,
                       alpha=0.9, edgecolors="white", linewidths=0.5,
                       label="Spec dışı ✗", zorder=5)
        if xs_miss:
            ax.scatter(xs_miss, ys_miss, zs_miss, c=TXT2, s=40,
                       alpha=0.5, edgecolors="#2a4060", linewidths=0.4,
                       label="Veri yok", zorder=3)

        # Optimum nokta
        if p.opt_solutions:
            best = p.opt_solutions[0]
            ox = best["factors"].get(fx["name"])
            oy = best["factors"].get(fy["name"])
            oz = best["factors"].get(fz["name"])
            if all(v is not None for v in [ox, oy, oz]):
                ax.scatter([ox], [oy], [oz], c=GOLD, s=200,
                           marker="*", zorder=6,
                           edgecolors="white", linewidths=0.8,
                           label=f"Optimum (D={best['desirability']:.3f})")

        ax.set_xlabel(fx["name"] + (f" ({fx['unit']})" if fx.get("unit") else ""),
                      color=TXT2, fontsize=8)
        ax.set_ylabel(fy["name"] + (f" ({fy['unit']})" if fy.get("unit") else ""),
                      color=TXT2, fontsize=8)
        ax.set_zlabel(fz["name"] + (f" ({fz['unit']})" if fz.get("unit") else ""),
                      color=TXT2, fontsize=8)
        ax.set_title("3D Faktör Uzayı — Run Noktaları", color=GOLD, fontsize=10)

        legend = ax.legend(fontsize=8, labelcolor=TXT,
                           facecolor=BG2, edgecolor="#2a4060",
                           loc="upper left")
        self.fig3d.tight_layout(pad=1.5)
        self.canvas3d.draw()



# ═══════════════════════════════════════════════════════════════════════════════
# SEKME 7 — DOĞRULAMA
# ═══════════════════════════════════════════════════════════════════════════════
class Tab7_Validation(QWidget):
    """Optimum formülasyon tahmin vs gerçek NGI karşılaştırması."""

    def __init__(self, project: OptimizerProject, app_ref, parent=None):
        super().__init__(parent)
        self.project = project
        self.app     = app_ref
        self._build()

    def _build(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(12, 12, 12, 12)
        outer.setSpacing(12)

        # ── Sol: Giriş paneli ─────────────────────────────────────────────────
        left = QVBoxLayout()
        left.setSpacing(10)

        # Optimum seçimi
        opt_card = card_frame()
        ol = QVBoxLayout(opt_card)
        ol.setContentsMargins(14, 10, 14, 10)
        ol.setSpacing(8)
        ol.addWidget(section_label("🏆  Doğrulanacak Formülasyon"))

        opt_row = QHBoxLayout()
        opt_row.addWidget(QLabel("Formülasyon:"))
        self.opt_combo = QComboBox(); self.opt_combo.setFixedHeight(28)
        self.opt_combo.currentIndexChanged.connect(self._on_opt_changed)
        opt_row.addWidget(self.opt_combo, 1)
        ol.addLayout(opt_row)

        # Seçili formülasyonun faktör değerleri
        self.factor_lbl = QLabel("")
        self.factor_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        self.factor_lbl.setWordWrap(True)
        ol.addWidget(self.factor_lbl)
        left.addWidget(opt_card)

        # Kabul kriteri
        crit_card = card_frame()
        cr = QHBoxLayout(crit_card)
        cr.setContentsMargins(14, 8, 14, 8)
        cr.setSpacing(16)
        cr.addWidget(QLabel("Kabul Kriteri (±%):"))
        self.crit_spin = QDoubleSpinBox()
        self.crit_spin.setRange(1.0, 30.0); self.crit_spin.setValue(15.0)
        self.crit_spin.setDecimals(1); self.crit_spin.setSingleStep(1.0)
        self.crit_spin.setFixedWidth(80); self.crit_spin.setFixedHeight(28)
        cr.addWidget(self.crit_spin)
        cr.addWidget(info_label("Tahmin ±% aralığında gerçek değer varsa PASS"))
        cr.addStretch()
        left.addWidget(crit_card)

        # Gerçek ölçüm girişi
        meas_card = card_frame()
        ml = QVBoxLayout(meas_card)
        ml.setContentsMargins(14, 10, 14, 10)
        ml.setSpacing(8)

        mhdr = QHBoxLayout()
        mhdr.addWidget(section_label("🔬  Gerçek NGI Ölçüm Sonuçları"))
        mhdr.addWidget(make_help_btn(
            "Optimum formülasyonu fiziksel olarak hazırlayıp NGI ölçümü yaptıktan sonra\n"
            "gerçek sonuçları buraya girin.\n\n"
            "Program tahmin ile gerçeği karşılaştırır ve model doğruluğunu raporlar.", self))
        mhdr.addStretch()
        ml.addLayout(mhdr)

        # Yanıt giriş tablosu
        self.meas_table = QTableWidget()
        self.meas_table.setColumnCount(4)
        self.meas_table.setHorizontalHeaderLabels(
            ["Yanıt", "Tahmin", "Gerçek Değer", "%95 PI"])
        self.meas_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.meas_table.setMinimumHeight(220)
        ml.addWidget(self.meas_table)

        # Giriş notu
        ml.addWidget(info_label(
            "💡  'Gerçek Değer' sütununa ölçüm sonuçlarınızı girin. "
            "Diğer sütunlar otomatik hesaplanır."))
        left.addWidget(meas_card, 1)

        # Değerlendir butonu
        self.btn_eval = make_btn("▶  Değerlendir", "rgba(20,80,20,0.9)", 42)
        self.btn_eval.setStyleSheet(self.btn_eval.styleSheet() +
                                    "font-size:14px; font-weight:bold;")
        self.btn_eval.clicked.connect(self._evaluate)
        left.addWidget(self.btn_eval)

        left_w = QWidget(); left_w.setLayout(left)
        left_w.setMinimumWidth(480); left_w.setMaximumWidth(560)

        # ── Sağ: Sonuç paneli ─────────────────────────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(8)

        # Genel skor
        self.score_card = card_frame()
        sc = QHBoxLayout(self.score_card)
        sc.setContentsMargins(20, 12, 20, 12)
        sc.setSpacing(30)

        self.lbl_overall  = QLabel("—")
        self.lbl_overall.setStyleSheet(
            f"color:{GOLD}; font-size:28px; font-weight:bold; background:transparent;")
        sc.addWidget(self.lbl_overall)

        self.lbl_pass_count = QLabel("—")
        self.lbl_pass_count.setStyleSheet(
            f"color:{TXT}; font-size:14px; background:transparent;")
        sc.addWidget(self.lbl_pass_count)

        self.lbl_model_quality = QLabel("—")
        self.lbl_model_quality.setStyleSheet(
            f"color:{TXT2}; font-size:12px; background:transparent;")
        self.lbl_model_quality.setWordWrap(True)
        sc.addWidget(self.lbl_model_quality, 1)
        right.addWidget(self.score_card)

        # Detay sonuç tablosu
        right.addWidget(section_label("📋  Tahmin vs Gerçek Karşılaştırma"))
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels(
            ["Yanıt", "Tahmin", "Gerçek", "Sapma %", "PI İçinde", "Karar"])
        self.result_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.result_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.result_table.setMinimumHeight(200)
        right.addWidget(self.result_table)

        # Grafik
        right.addWidget(section_label("📊  Görsel Karşılaştırma"))
        self.fig = Figure(figsize=(9, 3.5), facecolor=BG2)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setFixedHeight(220)
        self.canvas.setStyleSheet(f"background:{BG2};")
        right.addWidget(self.canvas)

        # Model güvenilirlik notu
        self.note_lbl = QLabel("")
        self.note_lbl.setStyleSheet(
            f"color:{TXT2}; font-size:11px; background:transparent;")
        self.note_lbl.setWordWrap(True)
        right.addWidget(self.note_lbl)

        right_w = QWidget(); right_w.setLayout(right)
        outer.addWidget(left_w)
        outer.addWidget(right_w, 1)

    # ── Public ───────────────────────────────────────────────────────────────
    def refresh(self):
        if not self.project.opt_solutions:
            return
        self._populate_opt_combo()
        self._build_meas_table()

    # ── İç metodlar ──────────────────────────────────────────────────────────
    def _populate_opt_combo(self):
        self.opt_combo.blockSignals(True)
        self.opt_combo.clear()
        for i, sol in enumerate(self.project.opt_solutions):
            d = sol["desirability"]
            self.opt_combo.addItem(f"#{i+1}  D={d:.4f}", i)
        self.opt_combo.blockSignals(False)
        self._on_opt_changed(0)

    def _on_opt_changed(self, idx):
        sol_idx = self.opt_combo.currentData()
        if sol_idx is None or sol_idx >= len(self.project.opt_solutions):
            return
        sol = self.project.opt_solutions[sol_idx]
        parts = []
        for f in self.project.factors:
            val = sol["factors"].get(f["name"], 0)
            unit = f.get("unit", "")
            parts.append(f"{f['name']}: {val:.4f} {unit}".strip())
        self.factor_lbl.setText("  |  ".join(parts))
        self._build_meas_table()

    def _get_current_sol(self):
        sol_idx = self.opt_combo.currentData()
        if sol_idx is None:
            return None
        sols = self.project.opt_solutions
        if sol_idx >= len(sols):
            return None
        return sols[sol_idx]

    def _build_meas_table(self):
        sol = self._get_current_sol()
        if sol is None:
            return

        active = [r for r in self.project.responses
                  if r in self.project.model_results]
        self.meas_table.setRowCount(len(active))

        for ri, resp_key in enumerate(active):
            label = RESPONSE_LABELS.get(resp_key, resp_key)
            pred  = sol["predictions"].get(resp_key)

            # %95 PI hesapla
            pi_lo, pi_hi = self._calc_pi(resp_key, sol)

            # Mevcut kayıtlı değeri yükle
            saved = self.project.validation_results.get(resp_key, {}).get("actual", "")

            items_data = [
                (label,    False),
                (f"{pred:.4f}" if pred is not None else "—", False),
                (str(saved) if saved != "" else "",           True),   # Düzenlenebilir
                (f"[{pi_lo:.3f}, {pi_hi:.3f}]"
                 if pi_lo is not None else "—",               False),
            ]
            for ci, (txt, editable) in enumerate(items_data):
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if editable:
                    item.setBackground(QColor("#0d2010"))
                    item.setForeground(QColor(TXT))
                else:
                    item.setBackground(QColor(BG3))
                    item.setForeground(QColor(TXT2))
                    item.setFlags(Qt.ItemFlag.ItemIsEnabled |
                                  Qt.ItemFlag.ItemIsSelectable)
                self.meas_table.setItem(ri, ci, item)

    def _calc_pi(self, resp_key, sol):
        """Tahmin için %95 güven aralığı (PI) hesapla."""
        try:
            import statsmodels.api as sm
            p      = self.project
            safe   = p.get_safe_names()
            factors = p.factors
            n_factors = len(factors)
            res    = p.model_results[resp_key]
            model  = res["model"]

            def encode(val, f):
                lo, hi = f["low"], f["high"]
                mid = f.get("mid", (lo+hi)/2)
                if hi == lo: return 0.0
                if val <= mid and (mid-lo) != 0:
                    return (val-lo)/(mid-lo) - 1
                elif (hi-mid) != 0:
                    return (val-mid)/(hi-mid)
                return 0.0

            row = {}
            for i, f in enumerate(factors):
                row[safe[i]] = encode(sol["factors"].get(f["name"], 0), f)
            for name in safe:
                row[f"{name}_sq"] = row[name]**2
            for a in range(n_factors):
                for b in range(a+1, n_factors):
                    row[f"{safe[a]}_x_{safe[b]}"] = row[safe[a]]*row[safe[b]]

            df_row  = pd.DataFrame([row])
            row_mat = sm.add_constant(df_row, has_constant="add")
            for col in model.params.index:
                if col not in row_mat.columns:
                    row_mat[col] = 0.0
            row_mat = row_mat.reindex(columns=model.params.index, fill_value=0.0)

            pred_res = model.get_prediction(row_mat)
            pi = pred_res.summary_frame(alpha=0.05)
            return float(pi["obs_ci_lower"].iloc[0]), float(pi["obs_ci_upper"].iloc[0])
        except Exception:
            return None, None

    def _evaluate(self):
        sol = self._get_current_sol()
        if sol is None:
            QMessageBox.warning(self, "", "Önce Optimizasyon sekmesinden optimum belirleyin.")
            return

        active = [r for r in self.project.responses
                  if r in self.project.model_results]

        # Gerçek değerleri topla
        measured = {}
        for ri, resp_key in enumerate(active):
            item = self.meas_table.item(ri, 2)
            if item is None:
                continue
            txt = item.text().strip().replace(",", ".")
            if txt:
                try:
                    measured[resp_key] = float(txt)
                except ValueError:
                    QMessageBox.warning(self, "Hata",
                        f"{RESPONSE_LABELS.get(resp_key,resp_key)}: geçersiz değer '{txt}'")
                    return

        if not measured:
            QMessageBox.warning(self, "", "En az 1 yanıt için gerçek değer girin.")
            return

        crit = self.crit_spin.value() / 100.0

        # Karşılaştırma hesapla
        comparisons = []
        for resp_key in active:
            if resp_key not in measured:
                continue
            pred   = sol["predictions"].get(resp_key)
            actual = measured[resp_key]
            pi_lo, pi_hi = self._calc_pi(resp_key, sol)

            if pred is not None and pred != 0:
                dev_pct = abs(actual - pred) / abs(pred) * 100
            else:
                dev_pct = None

            in_pi   = (pi_lo is not None and pi_hi is not None and
                       pi_lo <= actual <= pi_hi)
            in_crit = dev_pct is not None and dev_pct <= crit * 100
            passed  = in_pi or in_crit

            comparisons.append({
                "resp_key": resp_key,
                "pred":     pred,
                "actual":   actual,
                "dev_pct":  dev_pct,
                "in_pi":    in_pi,
                "passed":   passed,
                "pi_lo":    pi_lo,
                "pi_hi":    pi_hi,
            })

            # Sonuçları kaydet
            self.project.validation_results[resp_key] = {
                "actual":  actual,
                "passed":  passed,
                "dev_pct": dev_pct,
            }

        self._show_results(comparisons)

    def _show_results(self, comparisons):
        n_pass  = sum(1 for c in comparisons if c["passed"])
        n_total = len(comparisons)
        all_pass = n_pass == n_total

        # Genel skor
        score_color = GREEN if all_pass else GOLD if n_pass >= n_total*0.6 else RED
        score_text  = "PASS ✅" if all_pass else f"KISMİ ⚠  {n_pass}/{n_total}" \
                      if n_pass > 0 else "FAIL ✗"
        self.lbl_overall.setText(score_text)
        self.lbl_overall.setStyleSheet(
            f"color:{score_color}; font-size:22px; font-weight:bold; background:transparent;")
        self.lbl_pass_count.setText(
            f"{n_pass}/{n_total} yanıt\nkabul kriterini karşıladı")

        # Model güvenilirlik yorumu
        if all_pass:
            quality = "Model tahminleri doğrulandı. Formülasyon güvenle üretime alınabilir."
        elif n_pass >= n_total * 0.6:
            quality = ("Çoğu tahmin doğrulandı. Başarısız yanıtlar için modeli "
                       "ek run'larla iyileştirmeyi düşünün.")
        else:
            quality = ("Model yetersiz doğrulandı. Ek deney run'ları ekleyip "
                       "modeli yeniden kurun.")
        self.lbl_model_quality.setText(quality)

        # Detay tablo
        self.result_table.setRowCount(len(comparisons))
        for ri, c in enumerate(comparisons):
            label    = RESPONSE_LABELS.get(c["resp_key"], c["resp_key"])
            dev_str  = f"{c['dev_pct']:.2f}%" if c["dev_pct"] is not None else "—"
            pi_str   = "✅ Evet" if c["in_pi"] else "✗ Hayır"
            karar    = "PASS ✅" if c["passed"] else "FAIL ✗"
            bg_color = QColor("#0a2a0a") if c["passed"] else QColor("#2a0a0a")
            fg_karar = QColor(GREEN)     if c["passed"] else QColor(RED)

            row_items = [
                (label,                             TXT),
                (f"{c['pred']:.4f}" if c["pred"] is not None else "—", TXT2),
                (f"{c['actual']:.4f}",              TXT),
                (dev_str,
                 GREEN if c["dev_pct"] is not None and c["dev_pct"] <= 15 else
                 GOLD  if c["dev_pct"] is not None and c["dev_pct"] <= 25 else RED),
                (pi_str, GREEN if c["in_pi"] else RED),
                (karar, fg_karar),
            ]
            for ci, (txt, fg) in enumerate(row_items):
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                item.setBackground(bg_color)
                item.setForeground(fg if isinstance(fg, QColor) else QColor(fg))
                self.result_table.setItem(ri, ci, item)

        # Grafik
        self._plot_comparison(comparisons)

        # Öneri notu
        devs = [c["dev_pct"] for c in comparisons if c["dev_pct"] is not None]
        if devs:
            mean_dev = sum(devs) / len(devs)
            self.note_lbl.setText(
                f"Ortalama sapma: {mean_dev:.2f}%  |  "
                f"Maks sapma: {max(devs):.2f}%  |  "
                f"Model: {'Uygun ✅' if mean_dev <= 15 else 'Gözden Geçirin ⚠'}")

        self.app.status_bar.showMessage(
            f"Doğrulama tamamlandı — {n_pass}/{n_total} PASS", 6000)

    def _plot_comparison(self, comparisons):
        self.fig.clear()
        self.fig.patch.set_facecolor(BG2)

        n = len(comparisons)
        if n == 0:
            return

        axes = self.fig.subplots(1, 2)

        # Sol: Tahmin vs Gerçek scatter
        ax1 = axes[0]
        ax1.set_facecolor(BG3)
        ax1.tick_params(colors=TXT2, labelsize=8)
        for spine in ax1.spines.values():
            spine.set_edgecolor("#2a4060")

        preds   = [c["pred"] for c in comparisons if c["pred"] is not None]
        actuals = [c["actual"] for c in comparisons]
        colors  = [GREEN if c["passed"] else RED for c in comparisons]

        ax1.scatter(preds, actuals, c=colors, s=80, zorder=3,
                    edgecolors="white", linewidths=0.6, alpha=0.9)

        if preds and actuals:
            mn = min(min(preds), min(actuals))
            mx = max(max(preds), max(actuals))
            pad = (mx - mn) * 0.1 or 0.1
            ax1.plot([mn-pad, mx+pad], [mn-pad, mx+pad],
                     '--', color=GOLD, lw=1.2, label="İdeal")
            # ±15% bantları
            ax1.plot([mn-pad, mx+pad],
                     [(mn-pad)*0.85, (mx+pad)*0.85],
                     ':', color=TXT2, lw=0.8, alpha=0.5)
            ax1.plot([mn-pad, mx+pad],
                     [(mn-pad)*1.15, (mx+pad)*1.15],
                     ':', color=TXT2, lw=0.8, alpha=0.5, label="±15%")

        for c in comparisons:
            if c["pred"] is not None:
                ax1.annotate(
                    RESPONSE_LABELS.get(c["resp_key"], c["resp_key"])[:6],
                    (c["pred"], c["actual"]),
                    textcoords="offset points", xytext=(5, 3),
                    fontsize=7, color=TXT2)

        ax1.set_xlabel("Tahmin", color=TXT2, fontsize=8)
        ax1.set_ylabel("Gerçek", color=TXT2, fontsize=8)
        ax1.set_title("Tahmin vs Gerçek", color=GOLD, fontsize=9)
        ax1.legend(fontsize=7, labelcolor=TXT2,
                   facecolor=BG2, edgecolor="#2a4060")

        # Sağ: Sapma % çubuk grafik
        ax2 = axes[1]
        ax2.set_facecolor(BG3)
        ax2.tick_params(colors=TXT2, labelsize=8)
        for spine in ax2.spines.values():
            spine.set_edgecolor("#2a4060")

        labels  = [RESPONSE_LABELS.get(c["resp_key"], c["resp_key"])
                   for c in comparisons]
        devs    = [c["dev_pct"] if c["dev_pct"] is not None else 0
                   for c in comparisons]
        bar_col = [GREEN if d <= 15 else GOLD if d <= 25 else RED for d in devs]

        bars = ax2.bar(range(len(devs)), devs, color=bar_col,
                       alpha=0.85, edgecolor="#2a4060")
        ax2.set_xticks(range(len(labels)))
        ax2.set_xticklabels(labels, rotation=25, ha="right",
                             fontsize=7, color=TXT2)
        ax2.axhline(15, color=GOLD,  lw=1.2, linestyle="--", label="±15%")
        ax2.axhline(25, color=RED,   lw=1.0, linestyle=":",  label="±25%")
        ax2.set_ylabel("Sapma (%)", color=TXT2, fontsize=8)
        ax2.set_title("Tahmin Sapma %", color=GOLD, fontsize=9)
        ax2.legend(fontsize=7, labelcolor=TXT2,
                   facecolor=BG2, edgecolor="#2a4060")

        for bar, val in zip(bars, devs):
            ax2.text(bar.get_x() + bar.get_width()/2,
                     val + 0.3, f"{val:.1f}",
                     ha="center", va="bottom", color=TXT, fontsize=7)

        self.fig.tight_layout(pad=1.5)
        self.canvas.draw()


# ═══════════════════════════════════════════════════════════════════════════════
# PLACEHOLDER SEKMELER (Adım 5-7 için)
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# PDF RAPOR MODÜLÜ
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# PDF RAPOR MODÜLÜ v2 — Siyah/Beyaz, Türkçe Destekli
# ═══════════════════════════════════════════════════════════════════════════════
def generate_pdf_report(project, output_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, Image, HRFlowable, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io, datetime, os
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    # ── Türkçe Font Yükle ─────────────────────────────────────────────────────
    def load_fonts():
        """DejaVu fontlarını yükle — Türkçe karakter desteği."""
        font_paths = [
            # Linux
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            # Windows
            r"C:\Windows\Fonts\DejaVuSans.ttf",
            r"C:\Windows\Fonts\DejaVuSans-Bold.ttf",
            # PyInstaller bundle
            os.path.join(
                getattr(__import__('sys'), '_MEIPASS', ''),
                "DejaVuSans.ttf"),
        ]
        # DejaVu varsa yükle
        reg_path  = None
        bold_path = None
        for p in font_paths:
            if p and os.path.exists(p):
                if "Bold" in p:
                    bold_path = p
                else:
                    reg_path = p

        if reg_path:
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", reg_path))
                if bold_path:
                    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold_path))
                else:
                    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", reg_path))
                return "DejaVuSans", "DejaVuSans-Bold"
            except Exception:
                pass

        # Arial (Windows)
        arial_paths = [
            r"C:\Windows\Fonts\arial.ttf",
            r"C:\Windows\Fonts\arialbd.ttf",
        ]
        if os.path.exists(arial_paths[0]):
            try:
                pdfmetrics.registerFont(TTFont("ArialTR", arial_paths[0]))
                if os.path.exists(arial_paths[1]):
                    pdfmetrics.registerFont(TTFont("ArialTR-Bold", arial_paths[1]))
                else:
                    pdfmetrics.registerFont(TTFont("ArialTR-Bold", arial_paths[0]))
                return "ArialTR", "ArialTR-Bold"
            except Exception:
                pass

        # Fallback — Helvetica (Türkçe karakter yok ama en azından çalışır)
        return "Helvetica", "Helvetica-Bold"

    FONT_REG, FONT_BOLD = load_fonts()

    # ── Renkler (siyah/beyaz) ─────────────────────────────────────────────────
    BLACK   = colors.black
    WHITE   = colors.white
    DGRAY   = colors.HexColor("#333333")
    MGRAY   = colors.HexColor("#666666")
    LGRAY   = colors.HexColor("#cccccc")
    XLGRAY  = colors.HexColor("#f5f5f5")

    # ── Stiller ───────────────────────────────────────────────────────────────
    def S(name, font=None, size=10, color=BLACK,
          align=TA_LEFT, before=4, after=4, bold=False):
        return ParagraphStyle(
            name=name,
            fontName=FONT_BOLD if bold else (font or FONT_REG),
            fontSize=size,
            textColor=color,
            alignment=align,
            spaceBefore=before,
            spaceAfter=after,
        )

    s_title    = S("title",  size=20, align=TA_CENTER, bold=True,  before=20, after=6)
    s_subtitle = S("sub",    size=11, align=TA_CENTER, color=DGRAY, before=4, after=20)
    s_h1       = S("h1",     size=13, bold=True,  before=14, after=6)
    s_h2       = S("h2",     size=10, bold=True,  before=8,  after=4)
    s_body     = S("body",   size=9,  color=DGRAY, before=2, after=2)
    s_small    = S("small",  size=8,  color=MGRAY, before=1, after=1)
    s_center   = S("center", size=9,  align=TA_CENTER, before=2, after=2)
    s_verdict  = S("verdict",size=14, bold=True, align=TA_CENTER, before=6, after=6)

    # ── Tablo stili (siyah/beyaz, arka plan yok) ──────────────────────────────
    def tbl_style(has_header=True):
        cmds = [
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("FONTNAME",     (0,0), (-1,-1), FONT_REG),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("TEXTCOLOR",    (0,0), (-1,-1), BLACK),
            ("GRID",         (0,0), (-1,-1), 0.5, LGRAY),
            ("TOPPADDING",   (0,0), (-1,-1), 3),
            ("BOTTOMPADDING",(0,0), (-1,-1), 3),
            ("LEFTPADDING",  (0,0), (-1,-1), 5),
            ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ]
        if has_header:
            cmds += [
                ("FONTNAME",  (0,0), (-1,0), FONT_BOLD),
                ("FONTSIZE",  (0,0), (-1,0), 8),
                ("LINEBELOW", (0,0), (-1,0), 1.0, BLACK),
            ]
        return TableStyle(cmds)

    # ── Matplotlib fig → Image ────────────────────────────────────────────────
    def fig_to_img(fig, w_cm=15, h_cm=7):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=130,
                    bbox_inches="tight", facecolor="white")
        buf.seek(0)
        return Image(buf, width=w_cm*cm, height=h_cm*cm)

    # ── Kontur haritası üret ──────────────────────────────────────────────────
    def make_contour(resp_key):
        try:
            import statsmodels.api as sm
            p = project; safe = p.get_safe_names()
            factors = p.factors; n_f = len(factors)
            if n_f < 2 or resp_key not in p.model_results:
                return None
            xi, yi = 0, 1
            fx = factors[xi]; fy = factors[yi]
            n = 40
            x_a = np.linspace(fx["low"], fx["high"], n)
            y_a = np.linspace(fy["low"], fy["high"], n)
            XX, YY = np.meshgrid(x_a, y_a)
            ZZ = np.full_like(XX, np.nan)
            model = p.model_results[resp_key]["model"]
            param_idx = model.params.index

            def enc(v, f):
                lo, hi = f["low"], f["high"]
                mid = f.get("mid", (lo+hi)/2)
                if hi == lo: return 0.0
                if v <= mid and (mid-lo) != 0:
                    return (v-lo)/(mid-lo) - 1
                elif (hi-mid) != 0:
                    return (v-mid)/(hi-mid)
                return 0.0

            for r in range(n):
                for c in range(n):
                    row = {}
                    for k, f in enumerate(factors):
                        row[safe[k]] = enc(
                            XX[r,c] if k==xi else YY[r,c] if k==yi
                            else f.get("mid",(f["low"]+f["high"])/2), f)
                    for nm in safe:
                        row[f"{nm}_sq"] = row[nm]**2
                    for a in range(n_f):
                        for b in range(a+1, n_f):
                            row[f"{safe[a]}_x_{safe[b]}"] = row[safe[a]]*row[safe[b]]
                    df_r = pd.DataFrame([row])
                    rm = sm.add_constant(df_r, has_constant="add")
                    for col in param_idx:
                        if col not in rm.columns: rm[col] = 0.0
                    rm = rm.reindex(columns=param_idx, fill_value=0.0)
                    try: ZZ[r, c] = float(model.predict(rm)[0])
                    except: pass

            fig, ax = plt.subplots(figsize=(7, 4.5), facecolor="white")
            ax.set_facecolor("white")
            cf = ax.contourf(XX, YY, ZZ, levels=20, cmap="Blues")
            ax.contour(XX, YY, ZZ, levels=20, colors="gray",
                       alpha=0.4, linewidths=0.3)
            cb = fig.colorbar(cf, ax=ax)
            cb.ax.tick_params(labelsize=7)
            cb.set_label(RESPONSE_LABELS.get(resp_key, resp_key), fontsize=8)
            ax.set_xlabel(fx["name"] + (f" ({fx['unit']})" if fx.get("unit") else ""),
                          fontsize=8)
            ax.set_ylabel(fy["name"] + (f" ({fy['unit']})" if fy.get("unit") else ""),
                          fontsize=8)
            ax.set_title(RESPONSE_LABELS.get(resp_key, resp_key) + " Kontur Haritasi",
                         fontsize=9)
            ax.tick_params(labelsize=7)
            img = fig_to_img(fig, 14, 6)
            plt.close(fig)
            return img
        except Exception:
            return None

    # ── Desirability grafiği ──────────────────────────────────────────────────
    def make_desirability_img(solutions):
        try:
            fig, ax = plt.subplots(figsize=(8, 2.5), facecolor="white")
            ax.set_facecolor("white")
            labels = [f"#{i+1}" for i in range(len(solutions))]
            vals   = [s["desirability"] for s in solutions]
            bars   = ax.bar(labels, vals, color="steelblue", alpha=0.8,
                            edgecolor="black", linewidth=0.6)
            ax.set_ylim(0, 1.1)
            ax.axhline(0.8, color="green",  lw=1, linestyle="--", alpha=0.7)
            ax.axhline(0.5, color="orange", lw=1, linestyle="--", alpha=0.7)
            ax.set_ylabel("Desirability", fontsize=8)
            ax.set_title("Top-5 Desirability Skoru", fontsize=9)
            ax.tick_params(labelsize=8)
            for bar, val in zip(bars, vals):
                ax.text(bar.get_x() + bar.get_width()/2, val + 0.02,
                        f"{val:.3f}", ha="center", fontsize=8)
            img = fig_to_img(fig, 12, 4)
            plt.close(fig)
            return img
        except Exception:
            return None

    # ── Sayfa numaralandırma ──────────────────────────────────────────────────
    pn = [0]
    def on_page(cv, doc):
        pn[0] += 1
        cv.saveState()
        cv.setStrokeColor(LGRAY)
        cv.line(1.8*cm, 1.5*cm, A4[0]-1.8*cm, 1.5*cm)
        cv.setFont(FONT_REG, 8)
        cv.setFillColor(MGRAY)
        cv.drawString(1.8*cm, 0.8*cm, "Formulasyon-Optimizer  v1")
        cv.drawRightString(A4[0]-1.8*cm, 0.8*cm, f"Sayfa {pn[0]}")
        cv.drawCentredString(A4[0]/2, 0.8*cm,
                             datetime.datetime.now().strftime("%d.%m.%Y %H:%M"))
        cv.restoreState()

    # ── Belge ─────────────────────────────────────────────────────────────────
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=2.0*cm,  bottomMargin=2.2*cm,
    )
    W = A4[0] - 3.6*cm
    story = []

    def hr():
        return HRFlowable(width="100%", thickness=0.5,
                          color=BLACK, spaceAfter=8)

    def section(title, num):
        # Başlık bir sonraki sayfaya geçmesin — KeepTogether ile başlığı
        # içeriğin ilk elemanıyla birlikte tut
        story.append(KeepTogether([
            Spacer(1, 0.2*cm),
            Paragraph(f"{num}. {title}", s_h1),
            hr(),
        ]))

    def subsection_block(title, content_elements):
        """Alt başlık + içeriği aynı sayfada tut."""
        block = [Paragraph(title, s_h2)] + content_elements
        try:
            story.append(KeepTogether(block))
        except Exception:
            # KeepTogether çok büyük içerikte başarısız olursa düz ekle
            for el in block:
                story.append(el)

    # ════════════════════════════════════════
    # KAPAK
    # ════════════════════════════════════════
    story.append(Spacer(1, 2.5*cm))
    story.append(Paragraph("Formulasyon-Optimizer", s_title))
    story.append(Paragraph(
        "Farmasotik Inhaler Formülasyon Gelistirme ve Optimizasyon Raporu",
        s_subtitle))
    story.append(hr())

    dm  = project.design_matrix
    now = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    cov = [
        ["Parametre", "Deger"],
        ["Rapor Tarihi",  now],
        ["Tasarim Tipi",  project.design_type],
        ["Faktor Sayisi", str(len(project.factors))],
        ["Run Sayisi",    str(len(dm)) if dm is not None else "-"],
        ["Yanit Sayisi",  str(len(project.responses))],
        ["Model",         f"{len(project.model_results)} yanit modellendi"
                          if project.model_results else "Kurulmadi"],
        ["Optimizasyon",  f"{len(project.opt_solutions)} cozum bulundu"
                          if project.opt_solutions else "Yapilmadi"],
    ]
    ct = Table(cov, colWidths=[W*0.4, W*0.6])
    ct.setStyle(tbl_style())
    story.append(ct)
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 1. DENEY TASARIMI
    # ════════════════════════════════════════
    section("Deney Tasarimi", 1)

    fh = ["Faktor", "Birim", "Alt", "Orta", "Ust"]
    fd = [fh] + [
        [f["name"], f.get("unit","-"),
         f"{f['low']:.4g}",
         f"{f.get('mid',(f['low']+f['high'])/2):.4g}",
         f"{f['high']:.4g}"]
        for f in project.factors
    ]
    ft = Table(fd, colWidths=[W*0.30,W*0.12,W*0.19,W*0.19,W*0.20])
    ft.setStyle(tbl_style())
    story.append(KeepTogether([
        Paragraph("1.1 Faktorler", s_h2), ft, Spacer(1, 0.4*cm)
    ]))

    if dm is not None:
        story.append(Paragraph("1.2 Deney Matrisi", s_h2))
        mh = (["Run"] + [f["name"] for f in project.factors] +
              [RESPONSE_LABELS.get(r,r) for r in project.responses])
        md = [mh]
        for ri in range(len(dm)):
            row = [str(ri+1)]
            for f in project.factors:
                row.append(f"{dm.iloc[ri][f['name']]:.4g}")
            for resp in project.responses:
                v = project.run_results.get(ri,{}).get(resp,"")
                row.append(f"{v:.4g}" if isinstance(v, float) else "-")
            md.append(row)
        nc = len(mh)
        mt = Table(md, colWidths=[W/nc]*nc, repeatRows=1)
        mt.setStyle(tbl_style())
        story.append(mt)
    else:
        story.append(Paragraph("Deney matrisi olusturulmadi.", s_body))
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 2. MODEL
    # ════════════════════════════════════════
    section("Model Ozeti", 2)

    if not project.model_results:
        story.append(Paragraph("Model henuz kurulmadi.", s_body))
    else:
        story.append(KeepTogether([Paragraph("2.1 Model Istatistikleri", s_h2)]))
        story.append(Paragraph("", s_body))
        sh = ["Yanit","R2","Adj R2","RMSE","F p","SW p","n"]
        sd = [sh]
        for rk in project.responses:
            res = project.model_results.get(rk)
            if not res:
                sd.append([RESPONSE_LABELS.get(rk,rk),"HATA","-","-","-","-","-"])
                continue
            sw = res.get("sw_p")
            r2a = res['r2_adj']
            rmse = res['rmse']
            sd.append([
                RESPONSE_LABELS.get(rk,rk),
                f"{res['r2']:.4f}",
                f"{r2a:.4f}" if not (isinstance(r2a,float) and np.isnan(r2a)) else "nan",
                f"{rmse:.4f}" if not (isinstance(rmse,float) and np.isnan(rmse)) else "nan",
                f"{res['f_pvalue']:.4f}",
                f"{sw:.4f}" if sw else "-",
                str(res["n_obs"]),
            ])
        st2 = Table(sd, colWidths=[W*0.22,W*0.10,W*0.10,W*0.12,W*0.14,W*0.20,W*0.08],
                    repeatRows=1)
        st2.setStyle(tbl_style())
        story.append(st2)
        story.append(Spacer(1, 0.4*cm))

        for rk in project.responses:
            res = project.model_results.get(rk)
            if not res: continue
            story.append(Paragraph(
                f"2.2 Katsayilar — {RESPONSE_LABELS.get(rk,rk)}", s_h2))
            model = res["model"]
            cd = [["Terim","Katsayi","Std Hata","t","p"]]
            for nm in model.params.index:
                pv = model.pvalues[nm]
                cd.append([
                    nm,
                    f"{model.params[nm]:.4f}",
                    f"{model.bse[nm]:.4f}" if not np.isinf(model.bse[nm]) else "inf",
                    f"{model.tvalues[nm]:.3f}" if not np.isnan(model.tvalues[nm]) else "nan",
                    f"{pv:.4f}" if not np.isnan(pv) else "nan",
                ])
            ctt = Table(cd, colWidths=[W*0.35,W*0.16,W*0.16,W*0.16,W*0.17],
                        repeatRows=1)
            ctt.setStyle(tbl_style())
            story.append(ctt)
            story.append(Spacer(1, 0.3*cm))

            anova = res.get("anova")
            if anova is not None:
                story.append(Paragraph(
                    f"2.3 ANOVA — {RESPONSE_LABELS.get(rk,rk)}", s_h2))
                ad = [["Kaynak","SS","df","MS","p"]]
                for sn, row in anova.iterrows():
                    pv = row.get("PR(>F)", float("nan"))
                    ss = row.get("sum_sq", 0)
                    df_ = row.get("df", 1)
                    ad.append([
                        str(sn),
                        f"{ss:.4f}",
                        f"{int(df_)}",
                        f"{ss/max(df_,1):.4f}",
                        f"{pv:.4f}" if not np.isnan(pv) else "-",
                    ])
                att = Table(ad, colWidths=[W*0.35,W*0.16,W*0.10,W*0.16,W*0.17],
                            repeatRows=1)
                att.setStyle(tbl_style())
                story.append(att)
                story.append(Spacer(1, 0.3*cm))

    story.append(PageBreak())

    # ════════════════════════════════════════
    # 3. RESPONSE SURFACE
    # ════════════════════════════════════════
    section("Response Surface Grafikleri", 3)
    if not project.model_results:
        story.append(Paragraph("Model kurulmadi.", s_body))
    else:
        for i, rk in enumerate(project.responses):
            if rk not in project.model_results: continue
            story.append(Paragraph(
                f"3.{i+1}  {RESPONSE_LABELS.get(rk,rk)}", s_h2))
            img = make_contour(rk)
            if img:
                story.append(img)
            else:
                story.append(Paragraph("Grafik olusturulamadi.", s_small))
            story.append(Spacer(1, 0.3*cm))
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 4. OPTİMİZASYON
    # ════════════════════════════════════════
    section("Optimizasyon Sonuclari", 4)
    if project.spec_limits:
        story.append(KeepTogether([Paragraph("4.1 Spesifikasyon Sinirlari", s_h2)]))
        sph = ["Yanit","Hedef","LSL","USL","Hedef Deger","Agirlik s"]
        spd = [sph]
        for rk, sp in project.spec_limits.items():
            spd.append([
                RESPONSE_LABELS.get(rk,rk),
                sp.get("goal","-"),
                f"{sp['lsl']:.4g}" if sp.get("lsl") is not None else "-",
                f"{sp['usl']:.4g}" if sp.get("usl") is not None else "-",
                f"{sp['target']:.4g}" if sp.get("target") is not None else "-",
                f"{sp.get('weight',1.0):.1f}",
            ])
        spt = Table(spd, colWidths=[W*0.25,W*0.13,W*0.12,W*0.12,W*0.20,W*0.18])
        spt.setStyle(tbl_style())
        story.append(spt)
        story.append(Spacer(1, 0.4*cm))

    if project.opt_solutions:
        story.append(Paragraph("4.2 Top-5 Optimum Formülasyon", s_h2))
        sh2 = (["Sira","Desirability"] +
               [f["name"] for f in project.factors] +
               [RESPONSE_LABELS.get(r,r) for r in project.responses
                if r in project.model_results])
        sd2 = [sh2]
        for rank, sol in enumerate(project.opt_solutions):
            row = [f"#{rank+1}", f"{sol['desirability']:.4f}"]
            for f in project.factors:
                row.append(f"{sol['factors'].get(f['name'],0):.4g}")
            for rk in project.responses:
                if rk in project.model_results:
                    pv = sol["predictions"].get(rk)
                    row.append(f"{pv:.4f}" if pv is not None else "-")
            sd2.append(row)
        nc2 = len(sh2)
        st3 = Table(sd2, colWidths=[W/nc2]*nc2, repeatRows=1)
        st3.setStyle(tbl_style())
        story.append(st3)
        story.append(Spacer(1, 0.3*cm))
        di = make_desirability_img(project.opt_solutions)
        if di: story.append(di)
    else:
        story.append(Paragraph("Optimizasyon sonucu bulunamadi.", s_body))
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 5. DESIGN SPACE
    # ════════════════════════════════════════
    section("Design Space  (ICH Q8)", 5)
    if project.spec_limits and project.model_results:
        story.append(Paragraph(
            "Her CQA icin tanimlanan spesifikasyon sinirlari asagida verilmistir. "
            "Tasarim uzayi haritasi icin programin Design Space sekmesini kullaniniz.",
            s_body))
        story.append(Spacer(1, 0.3*cm))
        dsh = ["CQA","LSL","USL","Hedef"]
        dsd = [dsh]
        for rk, sp in project.spec_limits.items():
            dsd.append([
                RESPONSE_LABELS.get(rk,rk),
                f"{sp['lsl']:.4g}" if sp.get("lsl") is not None else "-",
                f"{sp['usl']:.4g}" if sp.get("usl") is not None else "-",
                sp.get("goal","-"),
            ])
        dst = Table(dsd, colWidths=[W*0.40,W*0.20,W*0.20,W*0.20])
        dst.setStyle(tbl_style())
        story.append(dst)
    else:
        story.append(Paragraph(
            "Design Space icin model ve spesifikasyon sinirlari gereklidir.", s_body))
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 6. DOĞRULAMA
    # ════════════════════════════════════════
    section("Dogrulama", 6)
    if not project.validation_results:
        story.append(Paragraph("Dogrulama henuz yapilmadi.", s_body))
    else:
        n_p = sum(1 for v in project.validation_results.values() if v.get("passed"))
        n_t = len(project.validation_results)
        verdict = f"SONUC: {'PASS' if n_p==n_t else 'KISMI'}  ({n_p}/{n_t})"
        story.append(Paragraph(verdict, s_verdict))
        story.append(Spacer(1, 0.3*cm))
        vh = ["Yanit","Tahmin","Gercek","Sapma %","Karar"]
        vd = [vh]
        for rk, vr in project.validation_results.items():
            pred = None
            if project.opt_solutions:
                pred = project.opt_solutions[0]["predictions"].get(rk)
            actual  = vr.get("actual","-")
            dev_pct = vr.get("dev_pct")
            passed  = vr.get("passed", False)
            vd.append([
                RESPONSE_LABELS.get(rk,rk),
                f"{pred:.4f}" if pred is not None else "-",
                f"{actual:.4f}" if isinstance(actual, float) else str(actual),
                f"{dev_pct:.2f}%" if dev_pct is not None else "-",
                "PASS" if passed else "FAIL",
            ])
        vt = Table(vd, colWidths=[W*0.30,W*0.17,W*0.17,W*0.17,W*0.19])
        vt.setStyle(tbl_style())
        story.append(vt)
    story.append(PageBreak())

    # ════════════════════════════════════════
    # 7. SONUÇ
    # ════════════════════════════════════════
    section("Sonuc ve Degerlendirme", 7)
    r2_vals = [res["r2"] for res in project.model_results.values()]
    if r2_vals:
        avg = sum(r2_vals)/len(r2_vals)
        me  = ("Model kalitesi mukemmel (ort. R2 >= 0.90)." if avg >= 0.9 else
               "Model kalitesi kabul edilebilir (ort. R2 >= 0.70)." if avg >= 0.7 else
               "Model kalitesi dusuk (ort. R2 < 0.70). Ek run onerilir.")
    else:
        me = "Model kurulmadi."

    if project.opt_solutions:
        bd = project.opt_solutions[0]["desirability"]
        oe = (f"Optimizasyon basarili (D={bd:.4f})." if bd >= 0.8 else
              f"Optimizasyon kismi basari (D={bd:.4f})." if bd >= 0.5 else
              f"Optimizasyon dusuk desirability (D={bd:.4f}). Gozden geciriniz.")
    else:
        oe = "Optimizasyon yapilmadi."

    if project.validation_results:
        np_ = sum(1 for v in project.validation_results.values() if v.get("passed"))
        nt_ = len(project.validation_results)
        ve  = (f"Dogrulama basarili ({np_}/{nt_} PASS)." if np_==nt_ else
               f"Dogrulama kismi ({np_}/{nt_} PASS).")
    else:
        ve = "Dogrulama yapilmadi."

    story.append(Paragraph(f"{me}  {oe}  {ve}", s_body))
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph("Kullanici Notlari:", s_h2))
    note_rows = [[""] for _ in range(5)]
    nt2 = Table(note_rows, colWidths=[W], rowHeights=[0.6*cm]*5)
    nt2.setStyle(TableStyle([
        ("GRID",    (0,0),(-1,-1), 0.5, LGRAY),
        ("FONTNAME",(0,0),(-1,-1), FONT_REG),
    ]))
    story.append(nt2)
    story.append(Spacer(1, 0.6*cm))
    sig = [
        ["Hazirlayan:", "Tarih:", "Imza:"],
        ["_"*28, "_"*18, "_"*28],
    ]
    sigt = Table(sig, colWidths=[W*0.40, W*0.25, W*0.35])
    sigt.setStyle(TableStyle([
        ("FONTNAME",  (0,0),(-1,0),  FONT_BOLD),
        ("FONTNAME",  (0,1),(-1,-1), FONT_REG),
        ("FONTSIZE",  (0,0),(-1,-1), 9),
        ("TEXTCOLOR", (0,0),(-1,-1), BLACK),
        ("ALIGN",     (0,0),(-1,-1), "LEFT"),
        ("TOPPADDING",(0,0),(-1,-1), 6),
    ]))
    story.append(sigt)

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)


class PlaceholderTab(QWidget):
    """Henüz geliştirilmemiş sekmeler için yer tutucu."""
    def __init__(self, title, description, icon="🔧", parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet(f"font-size: 48px; background: transparent;")
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(icon_lbl)

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(f"color: {GOLD}; font-size: 18px; font-weight: bold; background: transparent;")
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(title_lbl)

        desc_lbl = QLabel(description)
        desc_lbl.setStyleSheet(f"color: {TXT2}; font-size: 13px; background: transparent;")
        desc_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desc_lbl.setWordWrap(True)
        desc_lbl.setMaximumWidth(500)
        lay.addWidget(desc_lbl)

        dev_lbl = QLabel("— Geliştirme aşamasında —")
        dev_lbl.setStyleSheet(f"color: #3a5070; font-size: 11px; margin-top: 20px; background: transparent;")
        dev_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(dev_lbl)

    def refresh(self):
        pass


# ═══════════════════════════════════════════════════════════════════════════════
# ANA PENCERE
# ═══════════════════════════════════════════════════════════════════════════════
class FormulasyonOptimizerApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.project = OptimizerProject()
        self.setWindowTitle("Formulasyon-Optimizer  v1")
        self.setMinimumSize(1200, 750)
        self.resize(1400, 820)

        ico_path = resource_path("opt_icon.ico")
        if os.path.exists(ico_path):
            self.setWindowIcon(QIcon(ico_path))

        self._build_ui()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_lay = QVBoxLayout(central)
        main_lay.setContentsMargins(0, 0, 0, 0)
        main_lay.setSpacing(0)

        # ── Başlık çubuğu ────────────────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(60)
        header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0a1628,
                    stop:0.5 #0f2040,
                    stop:1 #0a1628
                );
                border-bottom: 2px solid {GOLD};
            }}
        """)
        hdr_lay = QHBoxLayout(header)
        hdr_lay.setContentsMargins(20, 0, 20, 0)

        # Sol: başlık
        title1 = QLabel("Formulasyon-Optimizer")
        title1.setStyleSheet(f"color: {GOLD}; font-size: 18px; font-weight: bold; background: transparent;")
        hdr_lay.addWidget(title1)

        sep = QLabel("│")
        sep.setStyleSheet(f"color: #2a4060; font-size: 20px; background: transparent;")
        hdr_lay.addWidget(sep)

        title2 = QLabel("Farmasötik İnhaler Formülasyon Geliştirme & Optimizasyon")
        title2.setStyleSheet(f"color: {TXT2}; font-size: 12px; background: transparent;")
        hdr_lay.addWidget(title2)
        hdr_lay.addStretch()

        # Sağ: versiyon + yeni proje butonu
        ver_lbl = QLabel(
            "v1.0  |  Ph.Eur 2.9.18 / USP <601>  "
            "|  Designed by Barbaros Çelik")
        ver_lbl.setStyleSheet(
            f"color: #3a5070; font-size: 10px; background: transparent;")
        hdr_lay.addWidget(ver_lbl)

        hdr_lay.addSpacing(16)
        btn_pdf = make_btn("📄  PDF Rapor", "rgba(20,80,20,0.5)", 30)
        btn_pdf.clicked.connect(self._export_pdf)
        hdr_lay.addWidget(btn_pdf)

        btn_tmpl = make_btn("📋  Excel Şablonu", "rgba(20,60,100,0.5)", 30)
        btn_tmpl.clicked.connect(self._export_excel_template)
        hdr_lay.addWidget(btn_tmpl)

        hdr_lay.addSpacing(8)
        btn_new = make_btn("🗋  Yeni Proje", "rgba(160,100,0,0.4)", 30)
        btn_new.clicked.connect(self._new_project)
        hdr_lay.addWidget(btn_new)

        main_lay.addWidget(header)

        # ── Sekmeler ─────────────────────────────────────────────────────────
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        main_lay.addWidget(self.tabs, 1)

        # Sekme 1 — Deney Tasarımı
        self.tab1 = Tab1_Design(self.project, self)
        self.tab1.design_generated.connect(self._on_design_generated)
        self.tabs.addTab(self.tab1, "1 · Deney Tasarımı")

        # Sekme 2 — Veri Girişi
        self.tab2 = Tab2_DataEntry(self.project, self)
        self.tabs.addTab(self.tab2, "2 · Veri Girişi")

        # Sekme 3 — Model
        self.tab3 = Tab3_Model(self.project, self)
        self.tab3.model_ready.connect(lambda: None)
        self.tabs.addTab(self.tab3, "3 · Model")

        # Sekme 4 — Response Surface
        self.tab4 = Tab4_ResponseSurface(self.project, self)
        self.tabs.addTab(self.tab4, "4 · Response Surface")

        # Sekme 5 — Optimizasyon
        self.tab5 = Tab5_Optimization(self.project, self)
        self.tabs.addTab(self.tab5, "5 · Optimizasyon")

        # Sekme 6 — Design Space
        self.tab6 = Tab6_DesignSpace(self.project, self)
        self.tabs.addTab(self.tab6, "6 · Design Space")

        # Sekme 7 — Doğrulama
        self.tab7 = Tab7_Validation(self.project, self)
        self.tabs.addTab(self.tab7, "7 · Doğrulama")

        # ── Durum çubuğu ─────────────────────────────────────────────────────
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Hazır — Deney Tasarımı sekmesinden başlayın.")

    def _on_tab_changed(self, idx):
        tab = self.tabs.widget(idx)
        if hasattr(tab, "refresh"):
            tab.refresh()

    def _on_design_generated(self):
        """Tasarım matrisi oluşturulduğunda diğer sekmeleri bilgilendir."""
        n = len(self.project.design_matrix) if self.project.design_matrix is not None else 0
        k = len(self.project.factors)
        self.status_bar.showMessage(
            f"Proje hazır — {n} run, {k} faktör, {len(self.project.responses)} yanıt  "
            f"|  Sonraki: Veri Girişi sekmesi", 8000)

    def _export_excel_template(self):
        """Veri girişi için Excel şablonu oluştur."""
        if self.project.design_matrix is None:
            QMessageBox.warning(self, "",
                "Önce Deney Tasarımı sekmesinden matris oluşturun.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Şablonu Kaydet",
            f"VeriGirisi_Sablonu_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import (
                PatternFill, Font, Alignment, Border, Side, Protection)
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Veri Girisi"

            # Renkler
            navy_fill  = PatternFill("solid", fgColor="002D62")
            gold_fill  = PatternFill("solid", fgColor="FFC600")
            green_fill = PatternFill("solid", fgColor="E2EFDA")
            gray_fill  = PatternFill("solid", fgColor="D9D9D9")
            blue_fill  = PatternFill("solid", fgColor="DDEEFF")
            white_fill = PatternFill("solid", fgColor="FFFFFF")

            thin = Side(style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def hdr_cell(ws, row, col, text, fill=navy_fill, font_color="FFFFFF",
                         bold=True, align="center"):
                c = ws.cell(row=row, column=col, value=text)
                c.fill = fill
                c.font = Font(bold=bold, color=font_color, size=10)
                c.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=True)
                c.border = border
                return c

            def data_cell(ws, row, col, value="", fill=white_fill, locked=False):
                c = ws.cell(row=row, column=col, value=value)
                c.fill = fill
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
                return c

            # ── Başlık ──────────────────────────────────────────────────────
            ws.merge_cells("A1:Z1")
            title_cell = ws["A1"]
            title_cell.value = "Formulasyon-Optimizer — Veri Giriş Şablonu"
            title_cell.fill = navy_fill
            title_cell.font = Font(bold=True, color="FFC600", size=13)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            # ── Bilgi satırı ─────────────────────────────────────────────────
            ws.merge_cells("A2:Z2")
            info_cell = ws["A2"]
            info_cell.value = (
                "KULLANIM: Mavi sütunlar = faktör değerleri (değiştirmeyin).  "
                "Yeşil sütunlar = NGI ölçüm sonuçları (bu sütunlara girin).  "
                "Doldurunca Veri Girişi sekmesinde 'Excel'den İçe Aktar' butonunu kullanın.")
            info_cell.fill = gold_fill
            info_cell.font = Font(bold=False, color="000000", size=9)
            info_cell.alignment = Alignment(horizontal="left", vertical="center",
                                            wrap_text=True)
            ws.row_dimensions[2].height = 30

            # ── Sütun başlıkları ─────────────────────────────────────────────
            dm = self.project.design_matrix
            factors = self.project.factors
            responses = self.project.responses

            col = 1
            # Run No
            hdr_cell(ws, 3, col, "Run No", fill=navy_fill)
            ws.column_dimensions[get_column_letter(col)].width = 8
            col += 1

            # Faktör sütunları — mavi
            for f in factors:
                unit = f.get("unit","")
                hdr_txt = (f["name"] + f" ({unit})" if unit else f["name"])
                hdr_cell(ws, 3, col, hdr_txt, fill=PatternFill("solid", fgColor="1F4E79"),
                         font_color="FFFFFF")
                ws.column_dimensions[get_column_letter(col)].width = 14
                col += 1

            # Yanıt sütunları — yeşil başlık
            for resp in responses:
                label = RESPONSE_LABELS.get(resp, resp)
                hdr_cell(ws, 3, col, label,
                         fill=PatternFill("solid", fgColor="375623"),
                         font_color="FFFFFF")
                ws.column_dimensions[get_column_letter(col)].width = 16
                col += 1

            ws.row_dimensions[3].height = 36

            # ── Veri satırları ───────────────────────────────────────────────
            for ri in range(len(dm)):
                row = ri + 4
                col = 1

                # Run No
                data_cell(ws, row, col, ri + 1,
                          fill=PatternFill("solid", fgColor="1F4E79"))
                ws.cell(row=row, column=col).font = Font(
                    bold=True, color="FFFFFF", size=9)
                col += 1

                # Faktör değerleri — salt okunur görünüm
                for f in factors:
                    val = dm.iloc[ri][f["name"]]
                    c = data_cell(ws, row, col,
                                  round(float(val), 6), fill=blue_fill)
                    col += 1

                # Yanıt sütunları — boş, girilecek
                for _ in responses:
                    data_cell(ws, row, col, "", fill=green_fill)
                    col += 1

                ws.row_dimensions[row].height = 18

            # ── Açıklama sayfası ─────────────────────────────────────────────
            ws2 = wb.create_sheet("Aciklama")
            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 60

            info_rows = [
                ("ŞABLON KULLANIM KILAVUZU", ""),
                ("", ""),
                ("Adım 1", "Bu şablonu indirin ve NGI deneylerini yapın."),
                ("Adım 2", "Yeşil sütunlara her run için NGI ölçüm sonuçlarını girin."),
                ("Adım 3", "Mavi faktör sütunlarını KESİNLİKLE değiştirmeyin."),
                ("Adım 4", "Programda 'Veri Girişi' sekmesine gidin."),
                ("Adım 5", "'Excel'den İçe Aktar' butonuna tıklayın ve bu dosyayı seçin."),
                ("Adım 6", "Veriler otomatik olarak yüklenecektir."),
                ("", ""),
                ("ÖNEMLİ NOTLAR", ""),
                ("Run sırası", "Satır sırasını değiştirmeyin — Run No kolonu referanstır."),
                ("Ondalık ayracı", "Virgül veya nokta kullanabilirsiniz, program her ikisini de okur."),
                ("Boş değerler", "Yapılamayan run'ları boş bırakabilirsiniz."),
                ("", ""),
                ("YANIT DEĞİŞKENLERİ", ""),
            ]

            for ri2, (rk, rl) in enumerate(RESPONSE_LABELS.items()):
                info_rows.append((rl, rk))

            for ri3, (label, value) in enumerate(info_rows, 1):
                c1 = ws2.cell(row=ri3, column=1, value=label)
                c2 = ws2.cell(row=ri3, column=2, value=value)
                if label in ("ŞABLON KULLANIM KILAVUZU", "ÖNEMLİ NOTLAR",
                             "YANIT DEĞİŞKENLERİ"):
                    c1.font = Font(bold=True, size=11, color="002D62")
                else:
                    c1.font = Font(bold=True, size=9)
                    c2.font = Font(size=9)
                c1.alignment = Alignment(vertical="center")
                c2.alignment = Alignment(vertical="center", wrap_text=True)
                ws2.row_dimensions[ri3].height = 16

            # ── Kaydet ───────────────────────────────────────────────────────
            wb.save(path)
            reply = QMessageBox.question(
                self, "Excel Şablonu Hazır",
                f"Excel şablonu oluşturuldu.\n\n"
                f"Şablonu şimdi açmak ister misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes)
            if reply == QMessageBox.StandardButton.Yes:
                import subprocess, platform
                if platform.system() == "Windows":
                    os.startfile(path)
                elif platform.system() == "Darwin":
                    subprocess.run(["open", path])
                else:
                    subprocess.run(["xdg-open", path])
            self.status_bar.showMessage(f"Excel şablonu kaydedildi: {path}", 5000)
        except Exception as e:
            import traceback
            write_log(f"Excel sablon hatasi:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Hata", f"Excel şablonu oluşturulamadı:\n{e}")

    def _export_pdf(self):
        """PDF rapor oluştur ve kaydet."""
        if self.project.design_matrix is None:
            QMessageBox.warning(self, "",
                "PDF rapor oluşturmak için önce deney matrisi oluşturun.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "PDF Rapor Kaydet",
            f"Formulasyon_Rapor_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            "PDF (*.pdf)")
        if not path:
            return
        try:
            self.status_bar.showMessage("PDF olusturuluyor...")
            QApplication.processEvents()
            generate_pdf_report(self.project, path)
            self.status_bar.showMessage(f"PDF kaydedildi: {path}", 6000)
            reply = QMessageBox.question(
                self, "PDF Rapor Hazir",
                f"PDF rapor basariyla olusturuldu.\n\n"
                f"Raporu simdi goruntülemek ister misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes)
            if reply == QMessageBox.StandardButton.Yes:
                import subprocess, platform
                if platform.system() == "Windows":
                    os.startfile(path)
                elif platform.system() == "Darwin":
                    subprocess.run(["open", path])
                else:
                    subprocess.run(["xdg-open", path])
        except Exception as e:
            import traceback
            write_log(f"PDF hata:\n{traceback.format_exc()}")
            QMessageBox.critical(self, "Hata",
                f"PDF olusturulamadi:\n{e}")

    def _new_project(self):
        reply = QMessageBox.question(
            self, "Yeni Proje",
            "Mevcut proje verisi silinecek. Devam edilsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.project.reset()
            self.tab1.refresh()
            self.tabs.setCurrentIndex(0)
            self.status_bar.showMessage("Yeni proje oluşturuldu.")


# ═══════════════════════════════════════════════════════════════════════════════
# GİRİŞ NOKTASI
# ═══════════════════════════════════════════════════════════════════════════════
def write_log(msg):
    """Hata logunu masaüstüne yazar."""
    import traceback, datetime
    try:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        log_path = os.path.join(desktop, "optimizer_log.txt")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"{datetime.datetime.now()}\n")
            f.write(msg + "\n")
    except Exception:
        pass

def exception_hook(exc_type, exc_value, exc_tb):
    import traceback
    msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    write_log(msg)
    try:
        QMessageBox.critical(None, "Hata",
            f"Beklenmedik hata oluştu:\n\n{exc_value}\n\n"
            f"Detaylar masaüstündeki optimizer_log.txt dosyasına kaydedildi.")
    except Exception:
        pass

def main():
    sys.excepthook = exception_hook
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(STYLE)

    # Qt exception handler
    def qt_msg_handler(mode, context, message):
        if "error" in message.lower() or "critical" in message.lower():
            write_log(f"Qt mesajı: {message}")
    from PyQt6.QtCore import qInstallMessageHandler
    qInstallMessageHandler(qt_msg_handler)

    window = FormulasyonOptimizerApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
